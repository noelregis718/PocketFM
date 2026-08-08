import openpyxl
from copy import copy

file_path = r'e:\Internship\PocketFM\Merged_Romance_Keywords.xlsx'
print(f"Loading {file_path}...")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

new_order = [
    1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, # Columns A to K (1 to 11) unchanged
    19, 20, 21,                        # Columns S, T, U (19, 20, 21) moved here
    12, 13, 14, 15, 16, 17, 18         # Columns L to R (12 to 18) containing Goodreads stats and others
]

print("Reordering columns while preserving formatting...")
ws_new = wb.create_sheet("Reordered")

for new_col_idx, old_col_idx in enumerate(new_order, start=1):
    old_letter = openpyxl.utils.get_column_letter(old_col_idx)
    new_letter = openpyxl.utils.get_column_letter(new_col_idx)
    
    # Copy width
    if old_letter in ws.column_dimensions:
        ws_new.column_dimensions[new_letter].width = ws.column_dimensions[old_letter].width

    for row_idx in range(1, ws.max_row + 1):
        old_cell = ws.cell(row=row_idx, column=old_col_idx)
        new_cell = ws_new.cell(row=row_idx, column=new_col_idx)
        
        new_cell.value = old_cell.value
        
        if old_cell.has_style:
            new_cell.font = copy(old_cell.font)
            new_cell.border = copy(old_cell.border)
            new_cell.fill = copy(old_cell.fill)
            new_cell.number_format = copy(old_cell.number_format)
            new_cell.alignment = copy(old_cell.alignment)

# Freeze top row
ws_new.freeze_panes = 'A2'

old_title = ws.title
wb.remove(ws)
ws_new.title = old_title

print("Saving workbook...")
wb.save(file_path)
print("Columns successfully rearranged.")
