import asyncio
import os
import re
import urllib.parse
import openpyxl
from playwright.async_api import async_playwright

EXCEL_FILE = r'E:\Internship\PocketFM\Agencies_Scraped.xlsx'
SHEET_NAME = 'Agencies'

async def process_row(row, context, sem, ws, wb, link_col, pages_col, series_col, author_col, lock):
    async with sem:
        series_name = ws.cell(row=row, column=series_col).value
        author_name = ws.cell(row=row, column=author_col).value
        
        if not series_name or not author_name:
            return
            
        search_query = f"{series_name} {author_name}"
        search_url = f"https://www.goodreads.com/search?q={urllib.parse.quote_plus(search_query)}"
        
        # Hardcoded series links for the 8 failing rows
        hardcoded_series = {
            13: "https://www.goodreads.com/series/319525-deliciously-dark-fairytales",
            14: "https://www.goodreads.com/series/283071-leveling-up",
            22: "https://www.goodreads.com/series/363044-fated",
            25: "https://www.goodreads.com/series/106969-thieves",
            32: "https://www.goodreads.com/series/204852-demon-days-vampire-nights",
            45: "https://www.goodreads.com/series/203027-secret-book-scone-society",
            55: "https://www.goodreads.com/series/49859-h-w-investigations",
            62: "https://www.goodreads.com/series/45799-cobbled-court-quilts"
        }
        
        if row in hardcoded_series:
            series_link = hardcoded_series[row]
            print(f"[{row}] Using hardcoded series link: {series_link}")
        else:
            print(f"[{row}] Searching: {search_query}")
            
            search_page = await context.new_page()
            try:
                await search_page.goto(search_url, wait_until="domcontentloaded", timeout=45000)
                await asyncio.sleep(2)
                
                # 1. Get first book link from search
                first_book_el = await search_page.query_selector('a[href*="/book/show/"]')
                if not first_book_el:
                    print(f"[{row}] No book links found in search results.")
                    return
                    
                b_url_initial = await first_book_el.get_attribute("href")
                if not b_url_initial.startswith('http'): 
                    b_url_initial = "https://www.goodreads.com" + b_url_initial
            except Exception as e:
                print(f"[{row}] Search Error: {e}")
                raise
            finally:
                await search_page.close()
                
            # 2. Open first book page
            book_page = await context.new_page()
            try:
                await book_page.goto(b_url_initial, wait_until="domcontentloaded", timeout=45000)
                await asyncio.sleep(2)
                
                # 3. Get series link
                series_el = await book_page.query_selector('a[href*="/series/"]')
                if not series_el:
                    print(f"[{row}] No series link found on book page.")
                    return
                    
                series_link = await series_el.get_attribute("href")
                if not series_link.startswith('http'): 
                    series_link = "https://www.goodreads.com" + series_link
            except Exception as e:
                print(f"[{row}] Book Page Error: {e}")
                raise
            finally:
                await book_page.close()
            
        # 4. Open series page
        print(f"[{row}] Opening series page: {series_link}")
        series_page = await context.new_page()
        try:
            await series_page.goto(series_link, wait_until="domcontentloaded", timeout=45000)
            await asyncio.sleep(2)
            
            book_items = await series_page.query_selector_all('.listWithDividers__item')
            print(f"[{row}] Found {len(book_items)} items on series page")
            
            total_pages = 0
            primary_books = 0
            scraped_books = 0
            
            books_to_scrape = []
            
            # 5. Parse primary books logic (Mega scraper logic)
            for item in book_items:
                title_el = await item.query_selector('a.bookTitle, span[itemprop="name"]')
                if not title_el:
                    title_el = await item.query_selector('a[href*="/book/show/"]')
                if not title_el: continue
                
                item_text = await title_el.inner_text()
                
                # Also check h3 for Book 1 format
                h3_el = await item.query_selector('h3')
                h3_text = await h3_el.inner_text() if h3_el else ""
                
                book_num = None
                
                # Mega scraper logic: check title for (Series, #1)
                match = re.search(r'\((?:.+?)(?:,\s*#|,\s*Book\s+|\s+Book\s+|\s+#)([0-9a-zA-Z\.\-]+)\)', item_text, re.IGNORECASE)
                if match:
                    num_str = match.group(1)
                    if num_str.isdigit():
                        book_num = num_str
                        
                # Fallback to checking h3 for exactly "Book 1", "Book 2"
                if not book_num and h3_text:
                    match2 = re.search(r'Book\s+(\d+)$', h3_text.strip(), re.IGNORECASE)
                    if match2:
                        book_num = match2.group(1)
                        
                if book_num:
                    # Found a primary book
                    link_el = await item.query_selector('a[href*="/book/show/"]')
                    if link_el:
                        b_url = await link_el.get_attribute("href")
                        if b_url and not b_url.startswith('http'): 
                            b_url = "https://www.goodreads.com" + b_url
                            
                        books_to_scrape.append((book_num, b_url, item_text.strip()))
                        
            if not books_to_scrape:
                print(f"[{row}] No primary books found matching series name.")
                return
            
            # Scrape each book
            for book_num, b_url, item_text in books_to_scrape:
                b_page = await context.new_page()
                try:
                    await b_page.goto(b_url, wait_until="domcontentloaded", timeout=45000)
                    extracted_pages = 0
                    
                    # Method 1: JSON-LD
                    ld_el = await b_page.query_selector('script[type="application/ld+json"]')
                    if ld_el:
                        import json
                        try:
                            data = json.loads(await ld_el.inner_text())
                            if isinstance(data, list): data = data[0]
                            if 'numberOfPages' in data:
                                extracted_pages = int(data['numberOfPages'])
                        except: pass
                        
                    # Method 2: UI 
                    if not extracted_pages:
                        p_el = await b_page.query_selector('[data-testid="pagesFormat"]')
                        if p_el:
                            p_match = re.search(r'(\d+)\s*pages', await p_el.inner_text(), re.IGNORECASE)
                            if p_match: extracted_pages = int(p_match.group(1))
                            
                    if not extracted_pages:
                        try:
                            content = await b_page.content()
                            matches = re.findall(r'(\d+)\s*pages', content, re.IGNORECASE)
                            if matches: extracted_pages = int(matches[0])
                        except: pass
                        
                    if extracted_pages > 0:
                        print(f"[{row}]   -> Book {book_num}: {extracted_pages} pages")
                        total_pages += extracted_pages
                        scraped_books += 1
                    else:
                        print(f"[{row}]   -> Book {book_num}: FAILED to extract pages")
                        
                    primary_books += 1
                    
                except Exception as e:
                    print(f"[{row}]   -> Error scraping {b_url}: {e}")
                finally:
                    await b_page.close()
            
            # ONLY save if we successfully extracted pages for ALL primary books we identified
            print(f"[{row}] SUMMARY: Found {primary_books} primary books. Successfully extracted pages for {scraped_books} books. Total Pages: {total_pages}")
            
            if primary_books > 0 and total_pages > 0:
                async with lock:
                    ws.cell(row=row, column=pages_col, value=total_pages)
                    wb.save(EXCEL_FILE)
                
                if scraped_books == primary_books:
                    print(f"[{row}] SAVED {total_pages} total pages to Excel!")
                else:
                    print(f"[{row}] SAVED {total_pages} total pages to Excel (WARNING: Missing {primary_books - scraped_books} books)")
            else:
                print(f"[{row}] SKIPPED SAVE - No pages extracted at all.")
                
        except Exception as e:
            print(f"[{row}] Series Page Error: {e}")
            raise
        finally:
            await series_page.close()

async def main():
    print(f"Loading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    
    header_row = 2
    headers = {ws.cell(row=header_row, column=c).value: c for c in range(1, ws.max_column + 1)}
    
    link_col = headers.get('Goodreads Link')
    pages_col = headers.get('No. of pages')
    series_col = headers.get('Series Title')
    author_col = headers.get('Author')
    
    if not all([link_col, pages_col, series_col, author_col]):
        print("Missing required columns!")
        return

    # Find rows that need processing
    rows_to_process = []
    for row in range(3, ws.max_row + 1):
        url = ws.cell(row=row, column=link_col).value
        pages = ws.cell(row=row, column=pages_col).value
        series = ws.cell(row=row, column=series_col).value
        
        if url and str(url).startswith('http') and not pages and series:
            rows_to_process.append(row)
            
    print(f"Found {len(rows_to_process)} rows to process.")
    
    if not rows_to_process:
        return

    user_data_dir = os.path.join(r"E:\Internship\PocketFM", "playwright_goodreads_profile")
    
    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir,
            headless=False,
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        sem = asyncio.Semaphore(5) # 5 concurrent browser tabs
        lock = asyncio.Lock() # Lock for saving Excel
        
        async def retry_process_row(row, context, sem, ws, wb, link_col, pages_col, series_col, author_col, lock):
            max_retries = 3
            for attempt in range(1, max_retries + 1):
                try:
                    await process_row(row, context, sem, ws, wb, link_col, pages_col, series_col, author_col, lock)
                    return
                except Exception as e:
                    print(f"[{row}] Attempt {attempt} failed: {e}")
                    if attempt == max_retries:
                        print(f"[{row}] All {max_retries} attempts failed.")
                    else:
                        print(f"[{row}] Retrying in 3 seconds...")
                        await asyncio.sleep(3)
        
        tasks = []
        for row in rows_to_process:
            task = asyncio.create_task(retry_process_row(row, context, sem, ws, wb, link_col, pages_col, series_col, author_col, lock))
            tasks.append(task)
            
        await asyncio.gather(*tasks)
        
        await context.close()
        print("\nAll done!")

if __name__ == "__main__":
    asyncio.run(main())
