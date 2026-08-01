import openpyxl
from openpyxl.styles import Alignment

filename = 'Mega_Sheet.xlsx'
print(f"Loading {filename}...")
wb = openpyxl.load_workbook(filename)
ws = wb.active

print("Applying text wrapping to all cells...")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        # Keep existing alignment properties if possible, but enable wrap_text
        current_horizontal = cell.alignment.horizontal if cell.alignment else None
        current_vertical = cell.alignment.vertical if cell.alignment else None
        cell.alignment = Alignment(horizontal=current_horizontal, vertical=current_vertical, wrap_text=True)

print("Enforcing thin row heights so they don't auto-expand...")
# Re-enforce thin row heights
for row_idx in range(1, ws.max_row + 1):
    ws.row_dimensions[row_idx].height = 12.5

print("Saving workbook...")
wb.save(filename)
print("Text wrapping applied successfully with thin rows enforced!")
