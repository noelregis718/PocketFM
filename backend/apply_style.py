import openpyxl

# 1. Set the path to your target Excel file
file_path = r'E:\Internship\PocketFM\Mega_Combined_List.xlsx'

try:
    print(f'Applying thin rows (15) and wrap text styling to {file_path}...')
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # 2. Iterate through all rows to set the height
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 15
        
        # 3. Iterate through every cell to apply text wrapping
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            
            # If the cell already has alignment properties, retain them but add wrapping
            if cell.alignment:
                cell.alignment = openpyxl.styles.Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    text_rotation=cell.alignment.text_rotation,
                    wrap_text=True,
                    shrink_to_fit=cell.alignment.shrink_to_fit,
                    indent=cell.alignment.indent
                )
            else:
                # Default alignment: Top-aligned and wrapped
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                
    # 4. Save the formatted workbook
    wb.save(file_path)
    print('Styling applied successfully.')
    
except Exception as e:
    print(f'Failed to apply styling: {e}')
