import docx
import openpyxl
from datetime import datetime

# 1. Parse Word Document for Title and Date
word_doc_path = r'e:\Internship\PocketFM\US Licensing Workstream Update.docx'
doc = docx.Document(word_doc_path)

# Table 5 seems to contain the "Signed Deals by Genre (YTD)"
target_table = doc.tables[5]
word_data = {}

for row_idx, row in enumerate(target_table.rows):
    if row_idx == 0: continue # Skip header
    cells = [cell.text.strip().replace('\n', ' ') for cell in row.cells]
    if len(cells) >= 7:
        title = cells[2].strip()
        date_str = cells[6].strip()
        if title and date_str:
            word_data[title.lower()] = (title, date_str)

print(f"Extracted {len(word_data)} titles from Word document.")

# 2. Parse Excel Master Sheet
excel_path = r'e:\Internship\PocketFM\US_Licensing_Lifecycle_Tracker.xlsx'
wb = openpyxl.load_workbook(excel_path)
ws_master = wb['Lifecycle Tracker - Master']

mismatches = [] # (Title, Word Date, Excel Date)

# Excel row order tracking
for row_idx, row in enumerate(ws_master.iter_rows(min_row=4), start=4):
    title_cell = row[4].value
    date_cell = row[13].value
    
    if not title_cell: continue
    title_excel = str(title_cell).strip()
    title_lower = title_excel.lower()
    
    if title_lower in word_data:
        word_title, word_date_str = word_data[title_lower]
        
        # Format Excel date to YYYY-MM-DD string for comparison
        excel_date_str = ""
        if isinstance(date_cell, datetime):
            excel_date_str = date_cell.strftime("%Y-%m-%d")
        elif date_cell:
            # Maybe it's already a string
            excel_date_str = str(date_cell).strip().split(' ')[0] # try to extract date part
            
        if word_date_str != excel_date_str:
            print(f"Mismatch found: {title_excel} | Word: {word_date_str} | Excel: {excel_date_str}")
            mismatches.append((title_excel, word_date_str, excel_date_str))

# 3. Create new sheet and write mismatches
if 'Date Mismatches' in wb.sheetnames:
    del wb['Date Mismatches']
    
ws_mismatches = wb.create_sheet('Date Mismatches')
ws_mismatches.append(['Book/Series Name', 'Incorrect Date (Word Doc)', 'Correct Date (Master Sheet)'])

for mismatch in mismatches:
    ws_mismatches.append(mismatch)

print(f"\nCreated 'Date Mismatches' sheet with {len(mismatches)} incorrect entries.")
wb.save(excel_path)
print("Done!")
