import openpyxl

file_path = r'e:\Internship\PocketFM\Merged_Romance_Keywords.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

def get_non_null(col_idx, limit=10):
    vals = set()
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col_idx).value
        if v:
            vals.add(str(v).strip())
            if len(vals) >= limit:
                break
    return list(vals)

print(f"Genre (4): {get_non_null(4)}")
print(f"Keyword (7): {get_non_null(7)}")
print(f"Genre Tags (13): {get_non_null(13)}")
print(f"Sub_Genre (18): {get_non_null(18)}")
