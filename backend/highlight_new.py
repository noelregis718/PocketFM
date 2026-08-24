import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

csv_path = r"E:\Internship\PocketFM\Romantasy _ Self Publication Master.csv"
excel_path = r"E:\Internship\PocketFM\Romantasy _ Self Publication Master_Highlighted.xlsx"

print("Loading CSV...")
df = pd.read_csv(csv_path)

wb = Workbook()
ws = wb.active
ws.title = "Master Data"

# Yellow fill for new cells
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

print("Writing to Excel and highlighting [NEW] entries...")
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Check if it's a string and contains [NEW]
        if isinstance(value, str) and "[NEW]" in value:
            # Clean the value by removing [NEW]
            clean_value = value.replace("[NEW]", "").strip()
            # Try to convert to float if it looks like a number (for lengths)
            try:
                if "." in clean_value:
                    cell.value = float(clean_value)
                else:
                    cell.value = int(clean_value)
            except ValueError:
                cell.value = clean_value
                
            # Highlight the cell!
            cell.fill = yellow_fill

# Freeze top row
ws.freeze_panes = 'A2'

print(f"Saving to {excel_path}...")
wb.save(excel_path)
print("Done! You can open the highlighted Excel file now.")
