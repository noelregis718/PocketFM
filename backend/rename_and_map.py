import openpyxl

file_path = r'e:\Internship\PocketFM\Merged_Romance_Keywords.xlsx'
print(f"Loading {file_path}...")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

col_Genre = 4
col_Keyword = 7

# Rename header
ws.cell(row=1, column=col_Genre).value = "Sub-Genre"

mapping = {
    "High Fantasy Court Adventure": ["Fae Court Romance", "Royal Fantasy Romance", "Fantasy Court Romance", "Kingdom Romance", "Royal Heir Romance", "Noble Fantasy Romance"],
    "Monster Romance (Non-Shifter)": ["Monster Romance", "Monster Boyfriend"],
    "High-Stakes Games & Deadly Trials": ["Fantasy Tournament Romance", "Progression Fantasy Romance", "Dungeon Trials Romance", "Deadly Trials", "Deadly Trials Romance"],
    "Paranormal Romance": ["Vampire Romance", "Witch Romance", "Demon Romance"],
    "Werewolf / Shifter Romance": ["Werewolf Romance", "Shifter Romance", "Fated Mates", "Werewolf and Shifter Romance"],
    "Mythology, Legend & Fairy Tale Retelling": ["Beauty and the Beast Retelling", "Hades and Persephone", "Greek Mythology Romance"],
    "Urban / Contemporary Fantasy Romance": ["Urban Fantasy Romance", "Hidden Magic Romance", "Urban Coven Romance", "Secret Magic World"],
    "Gothic Dark Romantasy": ["Gothic Romance", "Dark Fantasy Romance"],
    "War College / Military Academy": ["Magic Academy Romance", "Military Fantasy Romance", "War College Romance"],
    "Korean Romance Fantasy / Isekai": ["Villainess", "Isekai Romance", "LitRPG Romance"]
}

print("Applying sub-genre mapping to Column 4 (formerly Genre)...")
matched_count = 0

for row in range(2, ws.max_row + 1):
    kw = ws.cell(row=row, column=col_Keyword).value
    if kw:
        kw_str = str(kw).strip().lower()
        
        for subgenre, keywords in mapping.items():
            if any(k.lower() in kw_str for k in keywords):
                ws.cell(row=row, column=col_Genre).value = subgenre
                matched_count += 1
                break

print(f"Matched and updated {matched_count} rows in Column 4.")
print("Saving workbook...")
wb.save(file_path)
print("Mapping applied successfully.")
