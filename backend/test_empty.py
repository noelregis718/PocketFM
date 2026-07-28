import openpyxl

EXCEL_FILE = "Agency _ Publishers Crawl - 1852, Bent, Penzler, Biagi, AULit .xlsx"
SHEET_NAME = "1852 Literary Agent"
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb[SHEET_NAME]

empty_count = 0
for row in range(2, ws.max_row + 1):
    link = ws.cell(row=row, column=7).value
    pages = ws.cell(row=row, column=16).value
    if link and str(link).startswith("http") and (pages is None or str(pages).strip() == ""):
        print(f"Row {row}: Link={link}")
        empty_count += 1
print(f"Total left blank: {empty_count}")
