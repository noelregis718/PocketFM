import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_FILE = r"e:\Internship\PocketFM\noel_part2.xlsx"

def apply_styling(excel_file=EXCEL_FILE):
    print(f"Applying styling and wrapping to {excel_file}...")
    try:
        wb = load_workbook(excel_file)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Freeze the top row so headers stay visible when scrolling
            ws.freeze_panes = "A2"
            
            for col_idx, col in enumerate(ws.columns, 1):
                col_letter = get_column_letter(col_idx)
                
                # Make header bold
                header_cell = ws[f"{col_letter}1"]
                header_cell.font = Font(bold=True)
                
                # Set a wider standard width for readability
                ws.column_dimensions[col_letter].width = 35
                
            
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
                                 
            # Apply wrapping, top-alignment, and thin borders to all cells
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                # Lock row height to a thin size so it doesn't get massive from wrapping
                ws.row_dimensions[row[0].row].height = 20
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = thin_border

        wb.save(excel_file)
        print(f"Styling applied successfully! You can now open {excel_file}.")
    except Exception as e:
        print(f"Error styling excel: {e}")

if __name__ == "__main__":
    apply_styling()
