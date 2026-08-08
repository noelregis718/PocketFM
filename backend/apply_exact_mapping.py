import openpyxl

file_path = r'e:\Internship\PocketFM\Merged_Romance_Keywords.xlsx'
print(f"Loading {file_path}...")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

col_SubGenre = 4
col_Keyword = 7

mapping = {
    "High Fantasy Court Adventure ( Done )": ["Fae Court Romance", "Royal Fantasy Romance", "Fantasy Court Romance", "Kingdom Romance", "Royal Heir Romance", "Noble Fantasy Romance"],
    "High-Stakes Games & Deadly Trials (Done)": ["Fantasy Tournament Romance", "Progression Fantasy Romance", "Dungeon Trials Romance", "Deadly Trials"],
    "Paranormal Romance ( Done )": ["Vampire Romance", "Witch Romance", "Demon Romance"],
    "Werewolf / Shifter Romance ( Done )": ["Werewolf Romance", "Shifter Romance", "Fated Mates"],
    "Mythology, Legend & Fairy Tale Retelling ( Done )": ["Beauty and the Beast Retelling", "Hades and Persephone", "Greek Mythology Romance"]
}

print("Applying exact sub-genre mapping to Column 4...")
matched_count = 0

for row in range(2, ws.max_row + 1):
    kw = ws.cell(row=row, column=col_Keyword).value
    if kw:
        kw_str = str(kw).strip().lower()
        
        for subgenre, keywords in mapping.items():
            if any(k.lower() in kw_str for k in keywords):
                ws.cell(row=row, column=col_SubGenre).value = subgenre
                matched_count += 1
                break

print(f"Matched and updated {matched_count} rows in Column 4.")
print("Saving workbook...")
wb.save(file_path)
print("Exact mapping applied successfully.")
