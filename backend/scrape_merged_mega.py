import asyncio
import pandas as pd
import re
import os
import urllib.request
import urllib.parse
import json
from playwright.async_api import async_playwright

EXCEL_FILE = r"E:\Internship\PocketFM\Merged_Romance_Keywords.xlsx"
START_ROW = 4990
TARGET_ROWS = 5290
CONCURRENCY = 5
BATCH_SIZE = 50

# We use urllib for the Autocomplete API to bypass AWS WAF
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'application/json, text/javascript, */*; q=0.01',
}

def get_autocomplete_book_url(query):
    """Uses raw HTTP to bypass WAF and get the direct Book URL."""
    import time
    api_url = f"https://www.goodreads.com/book/auto_complete?format=json&q={urllib.parse.quote_plus(query)}"
    for attempt in range(3):
        try:
            req = urllib.request.Request(api_url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=15) as response:
                data = json.loads(response.read().decode("utf-8", errors="ignore"))
                if data and len(data) > 0:
                    book_path = data[0].get('bookUrl', '')
                    if book_path:
                        return "https://www.goodreads.com" + book_path
                return None
        except Exception as e:
            time.sleep(1)
    return None

async def process_row(index, row, df, context, sem):
    async with sem:
        book_name = str(row.get("Book Title", row.get("Series Name", ""))).strip()
        
        # SANITIZE BOOK NAME: Take only the part before a colon to avoid confusing the search
        if ":" in book_name:
            book_name = book_name.split(":")[0].strip()
            
        author_name = str(row.get("Author Name", "")).strip()
        
        if not book_name or book_name.lower() == 'nan':
            return
            
        print(f"\n--- Processing Row {index + 1} ---")

        # 1. Build Query
        if not author_name or author_name.lower() == 'nan' or author_name.lower() == 'none':
            print(f"[{index+2}] No valid author found. Querying: '{book_name}'")
            search_query = book_name
        else:
            print(f"[{index+2}] Querying Book + Author: '{book_name}' by '{author_name}'")
            search_query = f"{book_name} {author_name}"

        # 2. Bypass WAF using Autocomplete API
        book_url = await asyncio.to_thread(get_autocomplete_book_url, search_query)
        if not book_url:
            print(f"[{index+2}] Book not found in Autocomplete API. Falling back to Aggressive UI Search...")
            search_page = await context.new_page()
            try:
                search_url = f"https://www.goodreads.com/search?q={urllib.parse.quote_plus(search_query)}"
                await search_page.goto(search_url, wait_until="domcontentloaded", timeout=45000)
                await asyncio.sleep(2)
                
                b_link = await search_page.query_selector('a.bookTitle')
                if b_link:
                    book_url = await b_link.evaluate("el => el.href")
                    print(f"[{index+2}] Found Book URL via Aggressive Search: {book_url}")
                else:
                    print(f"[{index+2}] Still not found via Aggressive Search. Leaving completely blank for future retry.")
                    return
            except Exception as e:
                print(f"[{index+2}] Aggressive Search failed: {e}. Leaving completely blank for future retry.")
                return
            finally:
                await search_page.close()
        else:
            print(f"[{index+2}] Bypassed WAF. Found Book URL: {book_url}")
            
        df.at[index, "Goodreads Link"] = book_url
        
        # 3. Load Book Page using Playwright to render React
        page = await context.new_page()
        try:
            await page.goto(book_url, wait_until="domcontentloaded", timeout=45000)
            # Short sleep to allow React to inject data
            await asyncio.sleep(2)
            
            # EXPAND GENRES (Click "...more" if present)
            try:
                genre_btns = await page.query_selector_all('[data-testid="genresList"] button, [data-testid="genresList"] [role="button"]')
                for btn in genre_btns:
                    if "...more" in (await btn.inner_text()).lower():
                        await btn.click(force=True)
                        await asyncio.sleep(0.5)
                        break
            except Exception:
                pass

            # EXTRACT GENRES
            genres = []
            genre_els = await page.query_selector_all('[data-testid="genresList"] .Button__labelItem, .BookPageMetadataSection__genre a')
            for gel in genre_els:
                txt = (await gel.inner_text()).strip()
                if txt and txt not in genres: genres.append(txt)
            if genres:
                df.at[index, "Genre Tags"] = ", ".join(genres)
                
            # EXTRACT SYNOPSIS
            desc_el = await page.query_selector('[data-testid="description"] .Formatted, .readable')
            if desc_el: 
                df.at[index, "Synopsis"] = (await desc_el.inner_text()).strip()
                
            # EXTRACT RATINGS FOR STEP 4
            avg_rating = 0.0
            total_ratings = 0
            
            rating_el = await page.query_selector('div.RatingStatistics__rating')
            if rating_el:
                try: avg_rating = float((await rating_el.inner_text()).strip())
                except: pass
            
            count_el = await page.query_selector('[data-testid="ratingsCount"]')
            if count_el:
                try:
                    count_txt = (await count_el.inner_text()).replace(',', '').split()[0]
                    total_ratings = int(count_txt)
                except: pass
                
            df.at[index, "Goodreads Rating Book 1"] = avg_rating
            df.at[index, "Goodreads No. of Ratings Book 1"] = total_ratings
                
            # 4. Find Series Link dynamically
            series_tag = await page.query_selector('h3.Text__title3 a[href*="/series/"], [data-testid="series"] a, div.BookPageTitleSection__title a[href*="/series/"], a.infoBoxRowItem[href*="/series/"]')
            
            if not series_tag:
                # FALLBACK: Check if title implies it's a series but lacks a formal series page
                title_el = await page.query_selector('[data-testid="bookTitle"]')
                title_text = await title_el.inner_text() if title_el else ""
                
                # Matches "(Series Name Book 1)" or "(Series Name, #1)" etc.
                series_fallback_match = re.search(r'\((.+?)(?:,\s*#|,\s*Book\s+|\s+Book\s+|\s+#)\d+\)', title_text, re.IGNORECASE)
                
                if series_fallback_match:
                    fallback_series = series_fallback_match.group(1).strip()
                    print(f"[{index+2}] No formal Series Link, but detected informal Series in title: '{fallback_series}'. Running Fallback Series Scraping...")
                    df.at[index, "GoodReads_Series_URL"] = book_url # Use book 1 as the series URL fallback
                    
                    # Search for the series
                    fallback_search_url = f"https://www.goodreads.com/search?q={urllib.parse.quote_plus(fallback_series + ' ' + str(author_name))}"
                    search_page2 = await context.new_page()
                    try:
                        await search_page2.goto(fallback_search_url, wait_until="domcontentloaded", timeout=45000)
                        await asyncio.sleep(2)
                        
                        book_items = await search_page2.query_selector_all('tr[itemtype="http://schema.org/Book"]')
                        num_primary_books = 0
                        total_page_count = 0
                        
                        for item in book_items:
                            title_el_2 = await item.query_selector('a.bookTitle')
                            if not title_el_2: continue
                            item_text = await title_el_2.inner_text()
                            
                            if fallback_series.lower() not in item_text.lower():
                                continue
                                
                            match = re.search(r'\((?:.+?)(?:,\s*#|,\s*Book\s+|\s+Book\s+|\s+#)([0-9a-zA-Z\.\-]+)\)', item_text, re.IGNORECASE)
                            if match:
                                book_num = match.group(1)
                                if book_num.isdigit():
                                    num_primary_books += 1
                                    
                                    b_url_2 = await title_el_2.evaluate("el => el.href")
                                    b_page = await context.new_page()
                                    try:
                                        await b_page.goto(b_url_2, wait_until="domcontentloaded", timeout=30000)
                                        await asyncio.sleep(1)
                                        
                                        # Overwrite Book 1 definitive stats IF this is book 1
                                        if book_num == "1":
                                            df.at[index, "Goodreads Link"] = b_url_2
                                            b_rating_el = await b_page.query_selector('div.RatingStatistics__rating')
                                            if b_rating_el:
                                                try: df.at[index, "Goodreads Rating Book 1"] = float((await b_rating_el.inner_text()).strip())
                                                except: pass
                                                
                                            b_count_el = await b_page.query_selector('[data-testid="ratingsCount"]')
                                            if b_count_el:
                                                try:
                                                    b_count_txt = (await b_count_el.inner_text()).replace(',', '').split()[0]
                                                    df.at[index, "Goodreads No. of Ratings Book 1"] = int(b_count_txt)
                                                except: pass

                                        # Robust Page Extraction
                                        extracted_pages = 0
                                        try:
                                            ld_el = await b_page.query_selector('script[type="application/ld+json"]')
                                            if ld_el:
                                                import json
                                                data = json.loads(await ld_el.inner_text())
                                                if isinstance(data, list): data = data[0]
                                                if 'numberOfPages' in data:
                                                    extracted_pages = int(data['numberOfPages'])
                                        except: pass
                                        
                                        if not extracted_pages:
                                            p_el = await b_page.query_selector('[data-testid="pagesFormat"]')
                                            if p_el:
                                                p_match2 = re.search(r'(\d+)\s*pages', await p_el.inner_text(), re.IGNORECASE)
                                                if p_match2: extracted_pages = int(p_match2.group(1))
                                                
                                        if not extracted_pages:
                                            try:
                                                content = await b_page.content()
                                                matches = re.findall(r'(\d+)\s*pages', content, re.IGNORECASE)
                                                if matches: extracted_pages = int(matches[0])
                                            except: pass
                                            
                                        if extracted_pages:
                                            total_page_count += extracted_pages
                                            print(f"[{index+2}]   - Fallback Primary Book {book_num} | {extracted_pages} pages (fetched)")
                                        else:
                                            print(f"[{index+2}]   - Fallback Primary Book {book_num} | 0 pages (not found)")
                                    except Exception as e:
                                        pass
                                    finally:
                                        await b_page.close()
                        
                        if num_primary_books == 0:
                            num_primary_books = 1
                            
                        df.at[index, "Goodreads Primary Books Number"] = num_primary_books
                        df.at[index, "Goodreads Primary Books Page Count"] = total_page_count
                        print(f"[{index+2}] Extracted {num_primary_books} primary books | {total_page_count} total pages via Fallback Search.")
                    finally:
                        await search_page2.close()
                else:
                    print(f"[{index+2}] No Series link found on Book Page. It is likely a standalone.")
                    df.at[index, "GoodReads_Series_URL"] = book_url
                    df.at[index, "Goodreads Primary Books Number"] = 1
                    
                _apply_romantasy_checker(df, index, row, genres)
                return
                
            # 5. Extract Series URL
            series_url = await series_tag.evaluate("el => el.href")
            print(f"[{index+2}] Found Series Link via React: {series_url}")
            df.at[index, "GoodReads_Series_URL"] = series_url
            
            # 6. Navigate to series page to get primary book count and pages
            try:
                await page.goto(series_url, wait_until="domcontentloaded", timeout=45000)
                await asyncio.sleep(1)
                
                num_primary_books = 0
                total_page_count = 0
                
                book_items = await page.query_selector_all('.listWithDividers__item, .seriesWork')
                
                for item in book_items:
                    item_text = await item.inner_text()
                    
                    # Extract the exact book number string (e.g., '1', '1.5', '1-3', '1A')
                    match = re.search(r'Book\s+([0-9a-zA-Z\.\-]+)', item_text, re.IGNORECASE)
                    if match:
                        book_num = match.group(1)
                        # Check if it is purely a whole number (primary book)
                        if book_num.isdigit():
                            num_primary_books += 1
                            
                            # Aggressive scraping: Always visit the book page to get accurate page count
                            b_link = await item.query_selector('a.bookTitle, a[href*="/book/show"]')
                            if b_link:
                                b_url = await b_link.evaluate("el => el.href")
                                b_page = await context.new_page()
                                try:
                                    await b_page.goto(b_url, wait_until="domcontentloaded", timeout=30000)
                                    await asyncio.sleep(1)
                                    
                                    # DEFINITIVE BOOK 1 OVERRIDE
                                    if book_num == "1":
                                        df.at[index, "Goodreads Link"] = b_url
                                        b_rating_el = await b_page.query_selector('div.RatingStatistics__rating')
                                        if b_rating_el:
                                            try: df.at[index, "Goodreads Rating Book 1"] = float((await b_rating_el.inner_text()).strip())
                                            except: pass
                                            
                                        b_count_el = await b_page.query_selector('[data-testid="ratingsCount"]')
                                        if b_count_el:
                                            try:
                                                b_count_txt = (await b_count_el.inner_text()).replace(',', '').split()[0]
                                                df.at[index, "Goodreads No. of Ratings Book 1"] = int(b_count_txt)
                                            except: pass
                                    
                                    # Robust Page Extraction with Retry
                                    extracted_pages = 0
                                    
                                    for attempt in range(2):
                                        if extracted_pages > 0:
                                            break
                                            
                                        if attempt == 1:
                                            print(f"[{index+2}]   - Book {book_num} pages was 0. Rechecking (Attempt 2)...")
                                            await b_page.reload(wait_until="domcontentloaded")
                                            await asyncio.sleep(2)
                                            
                                        # 1. Try JSON-LD
                                        try:
                                            ld_el = await b_page.query_selector('script[type="application/ld+json"]')
                                            if ld_el:
                                                import json
                                                data = json.loads(await ld_el.inner_text())
                                                if isinstance(data, list): data = data[0]
                                                if 'numberOfPages' in data:
                                                    extracted_pages = int(data['numberOfPages'])
                                        except: pass
                                        
                                        # 2. Try simple data-testid
                                        if not extracted_pages:
                                            p_el = await b_page.query_selector('[data-testid="pagesFormat"]')
                                            if p_el:
                                                p_match2 = re.search(r'(\d+)\s*pages', await p_el.inner_text(), re.IGNORECASE)
                                                if p_match2: extracted_pages = int(p_match2.group(1))
                                        
                                        # 3. Try clicking Book details
                                        if not extracted_pages:
                                            try:
                                                btn = await b_page.query_selector('button:has-text("Book details")')
                                                if btn:
                                                    await btn.click(force=True)
                                                    await asyncio.sleep(1)
                                                    p_el = await b_page.query_selector('[data-testid="pagesFormat"]')
                                                    if p_el:
                                                        p_match2 = re.search(r'(\d+)\s*pages', await p_el.inner_text(), re.IGNORECASE)
                                                        if p_match2: extracted_pages = int(p_match2.group(1))
                                            except: pass
                                            
                                        # 4. Fallback Regex across entire content
                                        if not extracted_pages:
                                            try:
                                                content = await b_page.content()
                                                matches = re.findall(r'(\d+)\s*pages', content, re.IGNORECASE)
                                                if matches:
                                                    extracted_pages = int(matches[0])
                                            except: pass
    
                                    if extracted_pages:
                                        total_page_count += extracted_pages
                                        print(f"[{index+2}]   - Primary Book {book_num} | {extracted_pages} pages (fetched)")
                                    else:
                                        print(f"[{index+2}]   - Primary Book {book_num} | 0 pages (not found)")
                                except Exception as e:
                                    print(f"[{index+2}]   - Failed to scrape Book {book_num}: {e}")
                                finally:
                                    await b_page.close()
                            
                if num_primary_books == 0:
                    num_primary_books = max(1, len(book_items))
                    
                df.at[index, "Goodreads Primary Books Number"] = num_primary_books
                df.at[index, "Goodreads Primary Books Page Count"] = total_page_count
                print(f"[{index+2}] Extracted {num_primary_books} primary books | {total_page_count} total pages.")
            except Exception as e:
                print(f"[{index+2}] Error extracting primary books from series page: {e}")
            
            _apply_romantasy_checker(df, index, row, genres)

        except Exception as e:
            print(f"[{index+2}] Error rendering page via Playwright: {e}")
        finally:
            await page.close()

def _apply_romantasy_checker(df, index, row, genres):
    import pandas as pd
    classification = "Fail"
    all_tags = []
    
    def add_tags(val):
        if pd.isna(val) or not val: return
        parts = [p.strip() for p in str(val).split(',')]
        for p in parts:
            if p and p not in all_tags: all_tags.append(p)
            
    add_tags(row.get('Sub-Genre'))
    for g in genres: add_tags(g)
    add_tags(row.get('Keyword'))
    
    fantasy_kws = ['fantasy', 'paranormal', 'supernatural', 'magic', 'fae', 'witch', 'vampire', 'dragon', 'mythology', 'fairy tale', 'monster', 'isekai', 'reincarnation', 'sci-fi', 'beast']
    romance_kws = ['romance', 'romantasy', 'romantic', 'love', 'mate', 'heart', 'beauty']
    
    idx_f, idx_r = -1, -1
    for i, tag in enumerate(all_tags):
        tag_lower = tag.lower()
        if idx_f == -1 and any(kw in tag_lower for kw in fantasy_kws): idx_f = i
        if idx_r == -1 and any(kw in tag_lower for kw in romance_kws): idx_r = i
    for i, tag in enumerate(all_tags):
        if 'romantasy' in tag.lower():
            if idx_f == -1 or i < idx_f: idx_f = i
            if idx_r == -1 or i < idx_r: idx_r = i
            
    if idx_f != -1 and idx_r != -1:
        rank = max(idx_f, idx_r)
        if rank < 5: classification = "Strong Match"
        elif rank < 9: classification = "Confirmed Match"
        else: classification = "Weak Match"
        
    df.at[index, "Romantasy Checker"] = classification
    print(f"[{index+2}] Romantasy Checker: {classification}")

async def run_scraper():
    print(f"Loading {EXCEL_FILE}...")
    try:
        df = pd.read_excel(EXCEL_FILE)
    except Exception as e:
        print(f"Excel load error: {e}")
        return

    # Ensure all target columns exist
    gr_cols = ['GoodReads_Series_URL', 'Goodreads Primary Books Number', 'Goodreads Primary Books Page Count', 'Goodreads Rating Book 1', 'Goodreads No. of Ratings Book 1', 'Goodreads Link', 'Genre Tags', 'Romantasy Checker', 'Synopsis']
    for col in gr_cols:
        if col not in df.columns:
            df[col] = None

    print(f"Running MEGA-style Scraper on FIRST {TARGET_ROWS} ROWS with CONCURRENCY {CONCURRENCY}...")

    user_data_dir = os.path.join(r"E:\Internship\PocketFM", "playwright_goodreads_profile")
    
    total_to_process = min(TARGET_ROWS, len(df))
    
    for batch_start in range(START_ROW, total_to_process, BATCH_SIZE):
        batch_end = min(batch_start + BATCH_SIZE, total_to_process)
        print(f"\n=======================================================")
        print(f"STARTING BATCH {batch_start} to {batch_end} (Cooldown Architecture)")
        print(f"=======================================================\n")

        async with async_playwright() as p:
            print("Launching Persistent Playwright Context (Bypassing Captchas!)...")
            context = await p.chromium.launch_persistent_context(
                user_data_dir=user_data_dir,
                headless=False,
                args=['--disable-blink-features=AutomationControlled'],
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
                viewport={'width': 1280, 'height': 800}
            )
            
            sem = asyncio.Semaphore(CONCURRENCY)
            tasks = []
            for index in range(batch_start, batch_end):
                row = df.iloc[index]
                tasks.append(process_row(index, row, df, context, sem))
                
            await asyncio.gather(*tasks)
            await context.close()
            
        try:
            df.to_excel(EXCEL_FILE, index=False)
            print(f"\nBatch {batch_start}-{batch_end} saved successfully.")
        except Exception as e:
            print(f"Failed to save batch {batch_start}-{batch_end}: {e}")
            
    print("\nAll batches complete for the target rows.")
    excel_start = START_ROW + 2
    excel_end = START_ROW + TARGET_ROWS + 1
    apply_thin_row_styling(EXCEL_FILE, excel_start, excel_end)

def apply_thin_row_styling(file_path, start_row_idx, end_row_idx):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    print(f"Applying premium blue styling, column widths, and wrap text to rows {start_row_idx} to {end_row_idx}...")
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", name='Calibri', size=11)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alignment_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin_border = Border(left=Side(style='thin', color='BFBFBF'), 
                             right=Side(style='thin', color='BFBFBF'), 
                             top=Side(style='thin', color='BFBFBF'), 
                             bottom=Side(style='thin', color='BFBFBF'))
        regular_font = Font(name='Calibri', size=11)
        hyperlink_font = Font(color="0563C1", underline="single", name='Calibri', size=11)
        
        # Set specific column widths
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            cell_value = ws.cell(row=1, column=col).value
            header_str = str(cell_value).lower() if cell_value else ""
            
            if "synopsis" in header_str or "description" in header_str or "logline" in header_str:
                ws.column_dimensions[letter].width = 60
            elif "url" in header_str or "link" in header_str:
                ws.column_dimensions[letter].width = 30
            elif "name" in header_str or "title" in header_str or "author" in header_str or "keyword" in header_str:
                ws.column_dimensions[letter].width = 25
            else:
                ws.column_dimensions[letter].width = 15
        
        # Style Header
        ws.row_dimensions[1].height = 25
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = thin_border
            
        # Freeze panes if not already
        ws.freeze_panes = 'A2'
        
        max_r = min(end_row_idx, ws.max_row)
        
        for row in range(start_row_idx, max_r + 1):
            ws.row_dimensions[row].height = 17
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = alignment_left_top
                cell.border = thin_border
                
                # Identify links
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('http'):
                    cell.font = hyperlink_font
                    cell.hyperlink = cell.value
                else:
                    cell.font = regular_font
                    
        wb.save(file_path)
        print("Blue styling and wrapping applied successfully.")
    except Exception as e:
        print(f"Failed to apply styling: {e}")

if __name__ == "__main__":
    asyncio.run(run_scraper())
