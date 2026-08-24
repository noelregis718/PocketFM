import asyncio
import os
import urllib.parse
import openpyxl
from playwright.async_api import async_playwright

EXCEL_FILE = r'E:\Internship\PocketFM\Agencies_Scraped.xlsx'
SHEET_NAME = 'Agencies'

async def process_row(row, context, sem, ws, wb, link_col, series_col, author_col, lock):
    async with sem:
        series_name = ws.cell(row=row, column=series_col).value
        author_name = ws.cell(row=row, column=author_col).value
        
        if not series_name or not author_name:
            return
            
        hardcoded_series = {
            13: "https://www.goodreads.com/series/319525-deliciously-dark-fairytales",
            14: "https://www.goodreads.com/series/283071-leveling-up",
            22: "https://www.goodreads.com/series/363044-fated",
            25: "https://www.goodreads.com/series/106969-thieves",
            32: "https://www.goodreads.com/series/204852-demon-days-vampire-nights",
            45: "https://www.goodreads.com/series/203027-secret-book-scone-society",
            55: "https://www.goodreads.com/series/49859-h-w-investigations",
            62: "https://www.goodreads.com/series/45799-cobbled-court-quilts"
        }
        
        if row in hardcoded_series:
            series_link = hardcoded_series[row]
            print(f"[{row}] Using hardcoded Series Link: {series_link}")
            async with lock:
                ws.cell(row=row, column=link_col, value=series_link)
                wb.save(EXCEL_FILE)
            print(f"[{row}] Saved Series Link to Excel!")
            return
            
        search_query = f"{series_name} {author_name}"
        search_url = f"https://www.goodreads.com/search?q={urllib.parse.quote_plus(search_query)}"
        
        print(f"[{row}] Searching: {search_query}")
        
        search_page = await context.new_page()
        try:
            await search_page.goto(search_url, wait_until="domcontentloaded", timeout=45000)
            await asyncio.sleep(2)
            
            first_book_el = await search_page.query_selector('a[href*="/book/show/"]')
            if not first_book_el:
                print(f"[{row}] No book links found in search results.")
                return
                
            b_url_initial = await first_book_el.get_attribute("href")
            if not b_url_initial.startswith('http'): 
                b_url_initial = "https://www.goodreads.com" + b_url_initial
        except Exception as e:
            print(f"[{row}] Search Error: {e}")
            raise
        finally:
            await search_page.close()
            
        book_page = await context.new_page()
        try:
            await book_page.goto(b_url_initial, wait_until="domcontentloaded", timeout=45000)
            await asyncio.sleep(2)
            
            series_el = await book_page.query_selector('a[href*="/series/"]')
            if not series_el:
                print(f"[{row}] No series link found on book page.")
                return
                
            series_link = await series_el.get_attribute("href")
            if not series_link.startswith('http'): 
                series_link = "https://www.goodreads.com" + series_link
                
            print(f"[{row}] Found Series Link: {series_link}")
            
            async with lock:
                ws.cell(row=row, column=link_col, value=series_link)
                wb.save(EXCEL_FILE)
            print(f"[{row}] Saved Series Link to Excel!")
            
        except Exception as e:
            print(f"[{row}] Book Page Error: {e}")
            raise
        finally:
            await book_page.close()

async def main():
    print(f"Loading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    
    header_row = 2
    headers = {ws.cell(row=header_row, column=c).value: c for c in range(1, ws.max_column + 1)}
    
    link_col = headers.get('Goodreads Link')
    series_col = headers.get('Series Title')
    author_col = headers.get('Author')
    
    if not all([link_col, series_col, author_col]):
        print("Missing required columns!")
        return

    rows_to_process = []
    for row in range(3, ws.max_row + 1):
        series = ws.cell(row=row, column=series_col).value
        link = ws.cell(row=row, column=link_col).value
        if series and not link:
            rows_to_process.append(row)
            
    print(f"Found {len(rows_to_process)} rows to process.")
    
    if not rows_to_process:
        return

    user_data_dir = os.path.join(r"E:\Internship\PocketFM", "playwright_goodreads_profile")
    
    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir,
            headless=False,
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        sem = asyncio.Semaphore(5)
        lock = asyncio.Lock()
        
        async def retry_process_row(row, context, sem, ws, wb, link_col, series_col, author_col, lock):
            max_retries = 3
            for attempt in range(1, max_retries + 1):
                try:
                    await process_row(row, context, sem, ws, wb, link_col, series_col, author_col, lock)
                    return
                except Exception as e:
                    print(f"[{row}] Attempt {attempt} failed: {e}")
                    if attempt == max_retries:
                        print(f"[{row}] All {max_retries} attempts failed.")
                    else:
                        await asyncio.sleep(3)
        
        tasks = [asyncio.create_task(retry_process_row(row, context, sem, ws, wb, link_col, series_col, author_col, lock)) for row in rows_to_process]
        await asyncio.gather(*tasks)
        print("\nAll done!")

if __name__ == "__main__":
    asyncio.run(main())
