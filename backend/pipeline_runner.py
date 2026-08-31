import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from backend.scrapers.goodreads_series import get_primary_books_from_goodreads
from backend.scrapers.ocean_downloader import process_ocean_downloads
from backend.scrapers.zlib_scraper import process_zlib_downloads
from backend.utils.file_optimizer import process_file_optimizations
from backend.utils.pdf_converter import process_pdf_conversions
from urllib.parse import urlparse
from dotenv import load_dotenv

load_dotenv()

def sanitize_folder_name(name: str) -> str:
    """Removes invalid characters for folder names"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '')
    return name.strip()

import sys

def run_pipeline(start_row: int, end_row: int):
    excel_path = "e:\\Internship\\PocketFM\\CT _ US _ Pipeline Master Sheet new.xlsx"
    downloads_base = "e:\\Internship\\PocketFM\\downloads"
    
    if not os.path.exists(downloads_base):
        os.makedirs(downloads_base)
        
    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    
    if "Self-Pub Prioritization" not in wb.sheetnames:
        print("Error: 'Self-Pub Prioritization' sheet not found in the Excel file.")
        return
        
    ws = wb["Self-Pub Prioritization"]
    
    total_ocean_downloaded = 0
    total_zlib_downloaded = 0
    
    # Process rows in the specified range
    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row), start=start_row):
        series_name = row[4].value
        goodreads_link = row[5].value
        
        if not goodreads_link or not isinstance(goodreads_link, str):
            print(f"[Row {row_idx}] Skipping missing or invalid link.")
            continue
            
        if 'goodreads.com/series' not in goodreads_link:
            print(f"[Row {row_idx}] Skipping non-series link: {goodreads_link}")
            continue
            
        # 1. Scrape Goodreads for book list and true series name
        books, series_name = get_primary_books_from_goodreads(goodreads_link)
        if not books:
            print(f"[Row {row_idx}] No primary books found or failed to scrape series.")
            continue
            
        clean_series_name = sanitize_folder_name(series_name)
        series_dir = os.path.join(downloads_base, clean_series_name)
        
        if not os.path.exists(series_dir):
            os.makedirs(series_dir)
            
        print(f"\n{'='*50}")
        print(f"Processing Series: {clean_series_name}")
        print(f"Link: {goodreads_link}")
        print(f"Directory: {series_dir}")
        print(f"{'='*50}")
        
        # Pre-check: Skip books that already exist in the folder
        from backend.scrapers.ocean_downloader import sanitize_filename
        for book in books:
            safe_title_file = sanitize_filename(book.title)
            base_filename = f"{book.number}_{safe_title_file}"
            
            docx_path = os.path.join(series_dir, f"{base_filename}.docx")
            pdf_path = os.path.join(series_dir, f"{base_filename}.pdf")
            epub_path = os.path.join(series_dir, f"{base_filename}.epub")
            
            if os.path.exists(docx_path):
                print(f"[Smart Skip] {book.title} already exists as DOCX. Skipping download and conversion.")
                book.status = "completed"
                book.source = "Pre-existing"
            elif os.path.exists(pdf_path):
                print(f"[Smart Skip] {book.title} already exists as PDF. Skipping download, queuing for conversion.")
                book.status = "downloaded"
                book.pdf_path = pdf_path
                book.source = "Pre-existing"
            elif os.path.exists(epub_path):
                print(f"[Smart Skip] {book.title} already exists as EPUB. Skipping download, queuing for conversion.")
                book.status = "downloaded"
                book.epub_path = epub_path
                book.source = "Pre-existing"
            
        max_retries = 3
        for attempt in range(max_retries):
            # Check if there are books that still need to be processed (failed or not started)
            pending_books = [b for b in books if b.status != "completed"]
            if not pending_books:
                break
                
            if attempt > 0:
                print(f"\n[Row {row_idx}] Retrying {len(pending_books)} failed books (Attempt {attempt+1}/{max_retries})...")
                for b in pending_books:
                    b.status = "pending"
                    b.error_message = ""
                    # If it previously failed conversion, delete the bad PDF so it can be re-downloaded
                    if hasattr(b, 'pdf_path') and b.pdf_path and os.path.exists(b.pdf_path):
                        try:
                            os.remove(b.pdf_path)
                            b.pdf_path = None
                        except Exception:
                            pass
                            
            # 2. Download PDFs from OceanOfPDF
            process_ocean_downloads(books, series_dir)
            
            # 2.5 Fallback to Z-Library for failed downloads
            process_zlib_downloads(books, series_dir)
            
            # 3. Optimize files (compress large PDFs, convert EPUBs to DOCX)
            process_file_optimizations(books)
            
            # 4. Convert downloaded PDFs to Word docs
            process_pdf_conversions(books)
        
        success_count = sum(1 for b in books if b.status == "completed")
        failed_count = sum(1 for b in books if b.status == "conversion_failed")
        download_failed = sum(1 for b in books if b.status == "failed")
        
        ocean_count = sum(1 for b in books if getattr(b, 'source', None) == "OceanOfPDF" and b.status == "completed")
        zlib_count = sum(1 for b in books if getattr(b, 'source', None) == "Z-Library" and b.status == "completed")
        
        total_ocean_downloaded += ocean_count
        total_zlib_downloaded += zlib_count
        
        print(f"\n[Row {row_idx}] Finished processing series: {clean_series_name}")
        print(f"  - Successfully downloaded & converted to DOCX: {success_count} (OceanOfPDF: {ocean_count}, Z-Library: {zlib_count})")
        print(f"  - Failed DOCX conversion (PDF/EPUB kept): {failed_count}")
        print(f"  - Failed to download entirely: {download_failed}")
        
    print(f"\n{'='*50}")
    print("PIPELINE RUN COMPLETE")
    print(f"Total rows processed: {end_row - start_row + 1}")
    print(f"Grand Total - OceanOfPDF downloads: {total_ocean_downloaded}")
    print(f"Grand Total - Z-Library downloads:  {total_zlib_downloaded}")
    print(f"{'='*50}\n")
        
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Run the downloader pipeline for a range of rows.")
    parser.add_argument("--start", type=int, required=True, help="Starting row index (e.g. 216)")
    parser.add_argument("--end", type=int, required=True, help="Ending row index (e.g. 220)")
    args = parser.parse_args()
    
    run_pipeline(args.start, args.end)
