import openpyxl

file_path = r'e:\Internship\PocketFM\Merged_Romance_Keywords.xlsx'
print(f"Loading {file_path}...")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Target column: C (3)
# Source column: T (20)
col_C = 3
col_T = 20

print("Merging data from T into C...")
for row in range(2, ws.max_row + 1):
    val_C = ws.cell(row=row, column=col_C).value
    val_T = ws.cell(row=row, column=col_T).value
    # If C is empty and T has a value, copy T into C
    if (val_C is None or val_C == "") and val_T is not None:
        ws.cell(row=row, column=col_C).value = val_T

# Columns to delete: T (20) and S (19)
# We delete in descending order
cols_to_delete = [20, 19]

print("Deleting columns T and S...")
for col_idx in cols_to_delete:
    print(f"Deleting column {col_idx} ({ws.cell(row=1, column=col_idx).value})...")
    ws.delete_cols(col_idx)

print("Saving workbook...")
wb.save(file_path)
print("Data successfully merged and columns deleted.")
