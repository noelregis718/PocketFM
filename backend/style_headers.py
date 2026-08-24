import openpyxl
from openpyxl.styles import PatternFill, Font

file_path = r'E:\Internship\PocketFM\Mega_Combined_List.xlsx'

try:
    print(f'Applying blue headers to {file_path}...')
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Define a nice vibrant blue fill and white bold font for the headers
    blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    
    # Iterate through the first row (headers)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = blue_fill
        cell.font = white_font
                
    wb.save(file_path)
    print('Header styling applied successfully.')
    
except Exception as e:
    print(f'Failed to apply styling: {e}')
