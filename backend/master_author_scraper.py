import asyncio
import os
import sys
import pandas as pd
from playwright.async_api import async_playwright
from openpyxl import load_workbook

# Ensure backend folder is in path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from scraper import AuthorScraper

# Configuration
MASTER_FILE = r"E:\Internship\PocketFM\Romantasy _ Self Publication Master (1).xlsx"
OUTPUT_FILE = r"E:\Internship\PocketFM\Master_Author_Enrichment.xlsx"
SHEET_NAME = "Picks for Licensing"
HEADER_ROW = 2  # Row 2 in Excel
AUTHOR_COL_NAME = "Author Name"
EMAIL_COL_NAME = "Author Email ID"
WEBSITE_COL_NAME = "Author Contact Form - Website"
AGENCY_COL_NAME = "Agency Email ID"

MAX_CONCURRENT_TABS = 5
BATCH_SIZE = 20

async def process_author(context, author_name, author_scraper):
    """Fetch details for a single author."""
    if not author_name or str(author_name).lower() == "nan":
        return None
    
    print(f"  [Scrape] Processing Author: {author_name}...")
    details = await author_scraper.find_author_details(context, author_name)
    return details

async def run_master_author_mission():
    if not os.path.exists(MASTER_FILE):
        print(f"Error: {MASTER_FILE} not found.")
        return

    print(f"Loading Master File: {MASTER_FILE}")
    # Load with header=1 to get correct column names
    df = pd.read_excel(MASTER_FILE, sheet_name=SHEET_NAME, header=HEADER_ROW-1)
    
    # Identify rows needing enrichment
    # We enrich if any of the three target columns are empty
    mask = (
        df[EMAIL_COL_NAME].isna() | (df[EMAIL_COL_NAME].astype(str) == "N/A") |
        df[WEBSITE_COL_NAME].isna() | (df[WEBSITE_COL_NAME].astype(str) == "N/A") |
        df[AGENCY_COL_NAME].isna() | (df[AGENCY_COL_NAME].astype(str) == "N/A")
    )
    
    # Filter out rows where Author Name is missing
    mask &= df[AUTHOR_COL_NAME].notna()
    
    to_enrich_indices = df.index[mask].tolist()
    
    # FINAL FULL SWEEP: All remaining rows with missing data
    total_to_enrich = len(to_enrich_indices)
    
    if total_to_enrich == 0:
        print("Coverage is already 100% for author details. No enrichment needed.")
        return

    print(f"Locked onto: {total_to_enrich} authors for this mission batch.")
    
    author_scraper = AuthorScraper()
    semaphore = asyncio.Semaphore(MAX_CONCURRENT_TABS)

    async def sem_process(idx, delay):
        await asyncio.sleep(delay)
        async with semaphore:
            author_name = df.loc[idx, AUTHOR_COL_NAME]
            try:
                result = await asyncio.wait_for(
                    author_scraper.find_author_details(context, author_name),
                    timeout=120
                )
                if result:
                    df.at[idx, EMAIL_COL_NAME] = result.get("Author_Email", "N/A")
                    df.at[idx, WEBSITE_COL_NAME] = result.get("Contact_Website", result.get("Website", "N/A"))
                    df.at[idx, AGENCY_COL_NAME] = result.get("Agent_Email", "N/A")
            except Exception as e:
                print(f"  [Error] {author_name}: {e}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
        )
        
        # Load existing results to skip duplicates
        existing_authors = set()
        if os.path.exists(OUTPUT_FILE):
            existing_df = pd.read_excel(OUTPUT_FILE)
            existing_authors = set(existing_df[AUTHOR_COL_NAME].dropna().astype(str).tolist())
        
        # Process in chunks for saving
        for i in range(0, total_to_enrich, BATCH_SIZE):
            chunk_indices = to_enrich_indices[i : i + BATCH_SIZE]
            
            # Filter chunk to only include authors not already in results
            actual_indices = []
            for idx in chunk_indices:
                author_name = str(df.loc[idx, AUTHOR_COL_NAME]).strip()
                if author_name not in existing_authors:
                    actual_indices.append(idx)
            
            if not actual_indices:
                print(f"Batch {i//BATCH_SIZE + 1}: All authors already enriched. Skipping...")
                continue

            print(f"\n>>> Processing Batch {i//BATCH_SIZE + 1} ({len(actual_indices)} new authors)...")
            tasks = [sem_process(idx, count * 2) for count, idx in enumerate(actual_indices)]
            await asyncio.gather(*tasks)
            
            # Save progress to the SEPARATE OUTPUT FILE
            print(f"  [Progress] Saving batch {i//BATCH_SIZE + 1} to {OUTPUT_FILE}...")
            save_results_to_new_file(df, actual_indices)
            
            await asyncio.sleep(2) 

        await browser.close()
        print(f"\nMISSION ACCOMPLISHED. Results saved to {OUTPUT_FILE}")
        
        if os.name == 'nt':
            print(f"  [System] Opening results file: {OUTPUT_FILE}")
            os.startfile(OUTPUT_FILE)

def save_results_to_new_file(df, indices):
    """Saves the enriched rows by appending to the master results file."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    new_data = df.loc[indices, [AUTHOR_COL_NAME, EMAIL_COL_NAME, WEBSITE_COL_NAME, AGENCY_COL_NAME]]
    
    if os.path.exists(OUTPUT_FILE):
        try:
            # Load existing data and append
            existing_df = pd.read_excel(OUTPUT_FILE)
            combined_df = pd.concat([existing_df, new_data]).drop_duplicates(subset=[AUTHOR_COL_NAME], keep='last')
            combined_df.to_excel(OUTPUT_FILE, index=False)
        except Exception as e:
            print(f"  [Warning] Could not append to {OUTPUT_FILE}: {e}. Overwriting instead.")
            new_data.to_excel(OUTPUT_FILE, index=False)
    else:
        new_data.to_excel(OUTPUT_FILE, index=False)
    
    # Re-apply Styling to the whole file
    try:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        
        # 1. Header Styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 2. Zebra Striping & Content Alignment
        zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
        
        # 3. Freeze Panes (Lock Header)
        ws.freeze_panes = "A2"
        
        # 4. Column Auto-Width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            # Intelligent width calculation
            adjusted_width = (max_length + 2) * 1.1
            ws.column_dimensions[column].width = min(adjusted_width, 65)
            
        # 5. Filter
        ws.auto_filter.ref = ws.dimensions
        
        wb.save(OUTPUT_FILE)
    except Exception as e: 
        print(f"  [Warning] Styling failed: {e}")

if __name__ == "__main__":
    asyncio.run(run_master_author_mission())
