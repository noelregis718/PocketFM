import openpyxl

file_path = r'e:\Internship\PocketFM\Merged_Romance_Keywords.xlsx'
print(f"Loading {file_path}...")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

col_Keyword = 7
col_SubGenre = 18

mapping = {
    "High Fantasy Court Adventure": ["Fae Court Romance", "Royal Fantasy Romance", "Fantasy Court Romance", "Kingdom Romance", "Royal Heir Romance", "Noble Fantasy Romance"],
    "Monster Romance (Non-Shifter)": ["Monster Romance", "Monster Boyfriend"],
    "High-Stakes Games & Deadly Trials": ["Fantasy Tournament Romance", "Progression Fantasy Romance", "Dungeon Trials Romance", "Deadly Trials", "Deadly Trials Romance"],
    "Paranormal Romance": ["Vampire Romance", "Witch Romance", "Demon Romance"],
    "Werewolf / Shifter Romance": ["Werewolf Romance", "Shifter Romance", "Fated Mates", "Werewolf and Shifter Romance"],
    "Mythology, Legend & Fairy Tale Retelling": ["Beauty and the Beast Retelling", "Hades and Persephone", "Greek Mythology Romance"],
    "Urban / Contemporary Fantasy Romance": ["Urban Fantasy Romance", "Hidden Magic Romance", "Urban Coven Romance", "Secret Magic World"],
    "Gothic Dark Romantasy": ["Gothic Romance", "Dark Fantasy Romance"],
    "War College / Military Academy": ["Magic Academy Romance", "Military Fantasy Romance", "War College Romance"],
    "Korean Romance Fantasy / Isekai": ["Villainess", "Isekai Romance", "LitRPG Romance"]
}

print("Applying sub-genre mapping...")
matched_count = 0
unmatched_count = 0

for row in range(2, ws.max_row + 1):
    kw = ws.cell(row=row, column=col_Keyword).value
    if kw:
        kw_str = str(kw).strip().lower()
        found_match = False
        
        for subgenre, keywords in mapping.items():
            # Check if any mapped keyword is in the current cell's keyword
            if any(k.lower() in kw_str for k in keywords):
                ws.cell(row=row, column=col_SubGenre).value = subgenre
                found_match = True
                matched_count += 1
                break
                
        if not found_match:
            unmatched_count += 1

print(f"Matched {matched_count} rows. Unmatched: {unmatched_count} rows.")
print("Saving workbook...")
wb.save(file_path)
print("Mapping applied successfully.")
