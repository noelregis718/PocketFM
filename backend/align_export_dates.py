import docx
import openpyxl

print("Loading Word Document...")
word_doc_path = r'e:\Internship\PocketFM\US Licensing Workstream Update.docx'
doc = docx.Document(word_doc_path)
target_table = doc.tables[5]

# Extract Word data
word_data = {}
for row_idx, row in enumerate(target_table.rows):
    if row_idx == 0: continue
    cells = [cell.text.strip().replace('\n', ' ') for cell in row.cells]
    if len(cells) >= 7:
        title = cells[2].strip()
        date_str = cells[6].strip()
        if title and date_str:
            word_data[title.lower()] = date_str

print(f"Extracted {len(word_data)} titles from Word document.")

print("Loading Master Excel Sheet...")
excel_path = r'e:\Internship\PocketFM\US_Licensing_Lifecycle_Tracker.xlsx'
wb_master = openpyxl.load_workbook(excel_path, data_only=True) # data_only to grab existing values
ws_master = wb_master['Lifecycle Tracker - Master']

# Create new Excel workbook perfectly aligned
wb_new = openpyxl.Workbook()
ws_new = wb_new.active
ws_new.title = "Aligned Dates for Copy-Paste"

# To ensure perfect alignment for a direct copy-paste column, we will replicate the exact rows
for row_idx, row in enumerate(ws_master.iter_rows(min_row=1, max_row=ws_master.max_row), start=1):
    title_cell = row[4].value
    original_date_cell = row[13].value
    
    # We will write the Title in Col A, and the Date to copy in Col B
    title_str = str(title_cell) if title_cell is not None else ""
    date_out = original_date_cell # Default to keeping whatever was originally there
    
    # If this title is in our Word doc, OVERWRITE the date_out with the correct Word date
    if title_cell:
        title_lower = title_str.strip().lower()
        if title_lower in word_data:
            date_out = word_data[title_lower]
            
    # Write to new sheet exactly on the same row_idx
    # So if data starts on row 4 in Master, it will be row 4 here.
    ws_new.cell(row=row_idx, column=1, value=title_str)
    ws_new.cell(row=row_idx, column=2, value=date_out)

output_path = r'e:\Internship\PocketFM\Correct_Contract_Dates.xlsx'
wb_new.save(output_path)
print(f"Saved perfectly aligned sheet to {output_path}. You can now select Column B and paste it directly into the Master Sheet!")
