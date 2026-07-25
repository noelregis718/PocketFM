import asyncio
import urllib.request
from bs4 import BeautifulSoup
import re
import openpyxl
import random
import os

EXCEL_FILE = "Agency _ Publishers Crawl - 1852, Bent, Penzler, Biagi, AULit .xlsx"
SHEET_NAME = "1852 Literary Agent"
SERIES_LINK_COL = 7  # Column G
TOTAL_PAGES_COL = 16 # Column P

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9',
}

import urllib.request
import urllib.request
import urllib.error

def _fetch_html_sync(url):
    import time
    for attempt in range(3):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=60) as response:
                return response.read().decode("utf-8", errors="ignore")
        except urllib.error.HTTPError as e:
            if e.code in [403, 429, 503]:
                return -1 # Blocked
            return None
        except Exception as e:
            print(f"    Attempt {attempt+1} - Error fetching {url}: {e}")
            time.sleep(2)
    return None

async def fetch_html(url):
    return await asyncio.to_thread(_fetch_html_sync, url)

async def scrape_book_pages(book_url):
    html = await fetch_html(book_url)
    if html == -1: return -1
    if not html: return None
    
    soup = BeautifulSoup(html, "html.parser")
    
    # Try the modern React format container first
    pages_format = soup.find(attrs={"data-testid": "pagesFormat"})
    if pages_format:
        match = re.search(r'(\d+)\s+pages', pages_format.text)
        if match: return int(match.group(1))
        
    # Fallback to searching the entire text
    text = soup.get_text()
    match = re.search(r'(\d+)\s+pages', text)
    if match: return int(match.group(1))
    
    return None

async def process_series(series_url):
    html = await fetch_html(series_url)
    if html == -1: return -1
    if not html: return None
    
    soup = BeautifulSoup(html, "html.parser")
    primary_links = []
    
    # In raw Goodreads HTML, works are inside listWithDividers__item or seriesWork
    items = soup.find_all(class_="listWithDividers__item")
    if not items:
        items = soup.find_all(class_="seriesWork")
        
    for item in items:
        h3 = item.find("h3")
        if not h3: continue
        
        text = h3.get_text(strip=True)
        # Skip novellas/extras like "Book 1.5" or "Book 0.5"
        if re.search(r'Book\s+\d+\.\d+', text):
            continue
            
        # Treat as primary if it's an integer "Book 1" OR if it has no numbering at all (unnumbered series)
        a_tag = item.find("a", href=lambda href: href and "/book/show" in href)
        if a_tag:
            link = a_tag['href']
            if not link.startswith("http"):
                link = "https://www.goodreads.com" + link
            primary_links.append(link)
                
    if not primary_links:
        return None
        
    print(f"Found {len(primary_links)} primary books for {series_url}")
    
    total_pages = 0
    # Process up to 3 book pages in parallel per series to try and avoid bans
    semaphore = asyncio.Semaphore(10)
    
    async def bounded_scrape(url):
        async with semaphore:
            print(f"  -> Scraping book {url}")
            pages = await scrape_book_pages(url)
            await asyncio.sleep(random.uniform(0.5, 1.5))
            return url, pages

    results = await asyncio.gather(*(bounded_scrape(url) for url in primary_links))
    
    for url, pages in results:
        if pages == -1:
            return -1
        elif pages is None:
            print(f"  -> Failed to find page count for {url} (skipping this book but keeping total)")
            continue # The new rule: Skip missing page counts but keep the total of the rest!
            
        print(f"  -> Found {pages} pages")
        total_pages += pages
        
    return total_pages

async def main():
    print("Loading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    
    blocked = False
    
    for row in range(2, ws.max_row + 1):
        series_link = ws.cell(row=row, column=SERIES_LINK_COL).value
        current_pages = ws.cell(row=row, column=TOTAL_PAGES_COL).value
        
        # Skip empty series links or rows we already completed (unless they were marked as 0 by a previous bug)
        if not series_link or not str(series_link).startswith("http"):
            continue
        if current_pages is not None and str(current_pages).strip() != "" and str(current_pages).strip() != "0":
            continue
            
        print(f"\nRow {row}: Processing {series_link}")
        
        if "/book/show/" in series_link:
            print(f"  -> Standalone book URL detected. Scraping directly...")
            pages = await scrape_book_pages(series_link)
            if pages == -1:
                result = -1
            elif pages is None:
                result = None
            else:
                result = pages
        else:
            result = await process_series(series_link)
            
        if result == -1:
            print("!!! CAPTCHA OR BLOCK DETECTED. ABORTING SCRAPER !!!")
            blocked = True
            break
        elif result is None:
            print(f"Could not count all primary works. Leaving blank.")
        else:
            print(f"SUCCESS: Total primary pages = {result}")
            ws.cell(row=row, column=TOTAL_PAGES_COL, value=result)
            
        # Save instantly after every row
        wb.save(EXCEL_FILE)
        
        # Brief pause before next series row
        await asyncio.sleep(random.uniform(1.5, 3))
            
    print("Saving final Excel file...")
    wb.save(EXCEL_FILE)
    if blocked:
        print("Scraping aborted due to blocks, but progress was saved.")
    else:
        print("Scraping completed!")
if __name__ == "__main__":
    asyncio.run(main())
