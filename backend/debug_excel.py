from openpyxl import load_workbook
import os

file_path = r"E:\Internship\PocketFM\Romantasy _ Self Publication Master.xlsx"
wb = load_workbook(file_path, read_only=True)
ws = wb['Picks for Licensing']

print("Row 1:", [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))])
print("Row 2:", [cell for cell in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))])
print("Row 3:", [cell for cell in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))])
