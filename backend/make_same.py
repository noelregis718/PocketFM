import openpyxl

file_path = r'e:\Internship\PocketFM\Merged_Romance_Keywords.xlsx'
print(f"Loading {file_path}...")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

col_Genre = 4
col_Keyword = 7
col_GenreTags = 13
col_SubGenre = 18

print("Consolidating Genre, Keyword, Genre Tags, and Sub_Genre into Keyword (7) and Sub_Genre (18)...")
for row in range(2, ws.max_row + 1):
    c4 = ws.cell(row=row, column=col_Genre).value
    c7 = ws.cell(row=row, column=col_Keyword).value
    c13 = ws.cell(row=row, column=col_GenreTags).value
    c18 = ws.cell(row=row, column=col_SubGenre).value
    
    # Get the best value representing the keyword/genre
    val = c7 or c4 or c18 or c13
    
    ws.cell(row=row, column=col_Keyword).value = val
    ws.cell(row=row, column=col_SubGenre).value = val

print("Saving workbook...")
wb.save(file_path)
print("Sub-Genre and Keyword columns are now exactly the same for all rows.")
