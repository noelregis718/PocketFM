import asyncio
import os
import re
import urllib.parse
import openpyxl
from playwright.async_api import async_playwright

EXCEL_FILE = r'E:\Internship\PocketFM\Licensing _ Commissioning - Personal - Vikrant.xlsx'
SHEET_NAME = 'Agencies'

async def main():
    print(f"Loading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    
    # Find column indices (Header is on row 2)
    header_row = 2
    headers = {ws.cell(row=header_row, column=c).value: c for c in range(1, ws.max_column + 1)}
    
    link_col = headers.get('Goodreads Link')
    pages_col = headers.get('No. of pages')
    series_col = headers.get('Series Title')
    author_col = headers.get('Author')
    
    if not link_col or not pages_col:
        print("Could not find required columns!")
        return

    user_data_dir = os.path.join(r"E:\Internship\PocketFM", "playwright_goodreads_profile")
    
    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir,
            headless=False,
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        # We start from row 3
        for row in range(3, ws.max_row + 1):
            url = ws.cell(row=row, column=link_col).value
            pages_val = ws.cell(row=row, column=pages_col).value
            
            # Skip if already has pages, or doesn't have a URL
            if pages_val or not url or not str(url).startswith('http'):
                continue
                
            print(f"\n--- Processing Row {row} ---")
            print(f"URL: {url}")
            
            page = await context.new_page()
            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=45000)
                await asyncio.sleep(2)
                
                series_link = None
                total_pages = 0
                primary_books = 0
                
                series_el = await page.query_selector('a[href*="/series/"]')
                if series_el:
                    series_link = await series_el.get_attribute("href")
                    if not series_link.startswith('http'):
                        series_link = "https://www.goodreads.com" + series_link
                    print(f"  -> Found Series Link: {series_link}")
                
                if series_link:
                    series_page = await context.new_page()
                try:
                    await series_page.goto(series_link, wait_until="domcontentloaded", timeout=45000)
                    await asyncio.sleep(2)
                    
                    book_items = await series_page.query_selector_all('.listWithDividers__item')
                    print(f"  -> Found {len(book_items)} items in series list")
                    
                    for item in book_items:
                        h3_el = await item.query_selector('h3')
                        if not h3_el: continue
                        h3_text = await h3_el.inner_text()
                        
                        if not re.search(r'Book\s+\d+$', h3_text.strip(), re.IGNORECASE):
                            print(f"    [Skip] Non-primary: {h3_text.strip()}")
                            continue
                            
                        primary_books += 1
                        
                        book_link_el = await item.query_selector('a[href^="/book/show/"]')
                        if not book_link_el: 
                            print(f"    [Skip] Could not find link for {h3_text.strip()}")
                            continue
                            
                        b_url = await book_link_el.get_attribute("href")
                        if not b_url.startswith('http'): b_url = "https://www.goodreads.com" + b_url
                        
                        b_page = await context.new_page()
                        try:
                            await b_page.goto(b_url, wait_until="domcontentloaded", timeout=30000)
                            
                            extracted_pages = 0
                            ld_el = await b_page.query_selector('script[type="application/ld+json"]')
                            if ld_el:
                                import json
                                try:
                                    data = json.loads(await ld_el.inner_text())
                                    if isinstance(data, list): data = data[0]
                                    if 'numberOfPages' in data:
                                        extracted_pages = int(data['numberOfPages'])
                                except: pass
                                
                            if not extracted_pages:
                                p_el = await b_page.query_selector('[data-testid="pagesFormat"]')
                                if p_el:
                                    p_match2 = re.search(r'(\d+)\s*pages', await p_el.inner_text(), re.IGNORECASE)
                                    if p_match2: extracted_pages = int(p_match2.group(1))
                                    
                            print(f"    [+] {h3_text.strip()}: {extracted_pages} pages")
                            total_pages += extracted_pages
                        except Exception as e:
                            print(f"    [Error] {b_url}: {e}")
                        finally:
                            await b_page.close()
                            
                    print(f"  -> FINAL: {primary_books} primary books, {total_pages} total pages")
                    
                finally:
                    if series_link:
                        await series_page.close()
                        
                if total_pages == 0:
                    series_name = ws.cell(row=row, column=series_col).value
                    author_name = ws.cell(row=row, column=author_col).value
                    if series_name and author_name:
                        print(f"  -> Direct link yielded 0 pages. Running Fallback Search for '{series_name}' by '{author_name}'...")
                        
                        fallback_search_url = f"https://www.goodreads.com/search?q={urllib.parse.quote_plus(str(series_name) + ' ' + str(author_name))}"
                        search_page = await context.new_page()
                        try:
                            await search_page.goto(fallback_search_url, wait_until="domcontentloaded", timeout=45000)
                            await asyncio.sleep(2)
                            
                            book_items = await search_page.query_selector_all('tr[itemtype="http://schema.org/Book"]')
                            print(f"  -> Found {len(book_items)} items in fallback search")
                            
                            for item in book_items:
                                title_el_2 = await item.query_selector('a.bookTitle')
                                if not title_el_2: continue
                                item_text = await title_el_2.inner_text()
                                
                                if str(series_name).lower() not in item_text.lower():
                                    continue
                                    
                                match = re.search(r'\((?:.+?)(?:,\s*#|,\s*Book\s+|\s+Book\s+|\s+#)([0-9a-zA-Z\.\-]+)\)', item_text, re.IGNORECASE)
                                if match:
                                    book_num = match.group(1)
                                    if book_num.isdigit():
                                        b_url_2 = await title_el_2.get_attribute("href")
                                        if not b_url_2.startswith('http'): b_url_2 = "https://www.goodreads.com" + b_url_2
                                        
                                        b_page = await context.new_page()
                                        try:
                                            await b_page.goto(b_url_2, wait_until="domcontentloaded", timeout=30000)
                                            extracted_pages = 0
                                            ld_el = await b_page.query_selector('script[type="application/ld+json"]')
                                            if ld_el:
                                                import json
                                                try:
                                                    data = json.loads(await ld_el.inner_text())
                                                    if isinstance(data, list): data = data[0]
                                                    if 'numberOfPages' in data:
                                                        extracted_pages = int(data['numberOfPages'])
                                                except: pass
                                                
                                            if not extracted_pages:
                                                p_el = await b_page.query_selector('[data-testid="pagesFormat"]')
                                                if p_el:
                                                    p_match2 = re.search(r'(\d+)\s*pages', await p_el.inner_text(), re.IGNORECASE)
                                                    if p_match2: extracted_pages = int(p_match2.group(1))
                                                    
                                            print(f"    [Fallback] Book {book_num}: {extracted_pages} pages")
                                            total_pages += extracted_pages
                                            primary_books += 1
                                        except Exception as e:
                                            print(f"    [Fallback Error] {b_url_2}: {e}")
                                        finally:
                                            await b_page.close()
                            
                            print(f"  -> FALLBACK FINAL: {primary_books} primary books, {total_pages} total pages")
                        finally:
                            await search_page.close()
                    
                if total_pages > 0:
                    ws.cell(row=row, column=pages_col, value=total_pages)
                    wb.save(EXCEL_FILE)
                    print(f"  -> Saved to Excel!")
                    
            except Exception as e:
                print(f"  -> Main Error: {e}")
            finally:
                await page.close()
                
        print("\nAll done!")

if __name__ == "__main__":
    asyncio.run(main())
