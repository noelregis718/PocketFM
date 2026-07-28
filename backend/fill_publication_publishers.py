import asyncio
import os
import openpyxl
import random
import re
from playwright.async_api import async_playwright
import sys

# Add current directory to path to import GoodreadsScraper if needed
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from goodreads_scraper import GoodreadsScraper

INPUT_FILE = r"e:\Internship\PocketFM\Publication Master Sheet.xlsx"

class BlockedError(Exception):
    pass

async def solve_captcha_if_present(page, url=""):
    try:
        page_text = await page.evaluate("document.body ? document.body.innerText.toLowerCase() : ''")
        if "unexpected error" in page_text or "403 forbidden" in page_text or ("captcha" in page_text and not await page.query_selector('.bookTitle')):
            if await page.query_selector('#captcha-image, .captcha, iframe[src*="captcha"]'):
                print(f"    [!!!] CAPTCHA detected! Please solve it manually.")
                try:
                    await page.wait_for_selector('a.bookTitle, [data-testid="pagesFormat"], .listWithDividers__item, h1', timeout=120000)
                    print(f"    [Success] CAPTCHA solved.")
                    return
                except:
                    print(f"    [Timeout] CAPTCHA wait timeout. Blocked.")
                    raise BlockedError("CAPTCHA timeout")
            else:
                print(f"    [!!!] Unexpected Goodreads error page detected. Blocked.")
                raise BlockedError("Unexpected error page")
    except BlockedError:
        raise
    except Exception:
        pass

async def process_row(row_idx, ws, col_map, context, semaphore, excel_lock, wb):
    series_url = ws.cell(row=row_idx, column=col_map['Goodreads Series URL']).value
    
    if not series_url or not str(series_url).startswith("http"):
        print(f"  [Row {row_idx}] Invalid or empty Goodreads Series URL.")
        return
        
    async with semaphore:
        print(f"\n[Row {row_idx}] Processing Series: {series_url}")
        page = await context.new_page()
        try:
            # 1. Visit Series URL
            try:
                await page.goto(series_url, wait_until="domcontentloaded", timeout=45000)
            except Exception as e:
                if "Timeout" not in str(e): raise
            await asyncio.sleep(random.uniform(1.5, 3.5))
            await solve_captcha_if_present(page, series_url)
            
            # 2. Get first book URL
            book_links = await page.query_selector_all('.listWithDividers__item a.bookTitle, .listWithDividers__item a[href*="/book/show/"]')
            if not book_links:
                print(f"  [Row {row_idx}] Could not find any books in series.")
                return
                
            first_book_href = await book_links[0].evaluate("el => el.href")
            if first_book_href.startswith('/'):
                first_book_href = f"https://www.goodreads.com{first_book_href}"
                
            print(f"  [Row {row_idx}] Navigating to First Book: {first_book_href}")
            
            # 3. Visit First Book URL
            try:
                await page.goto(first_book_href, wait_until="domcontentloaded", timeout=45000)
            except Exception as e:
                if "Timeout" not in str(e): raise
            await asyncio.sleep(random.uniform(1.5, 3.5))
            await solve_captcha_if_present(page, first_book_href)
            
            # Scroll down to trigger lazy loading of the Book Details section
            print(f"  [Row {row_idx}] Scrolling down to load Book Details...")
            await page.evaluate("window.scrollBy(0, 800)")
            await asyncio.sleep(1)
            await page.evaluate("window.scrollBy(0, 800)")
            await asyncio.sleep(1)
            await page.keyboard.press("PageDown")
            await asyncio.sleep(1.5)
            
            # Click the "Book Details and Editions" dropdown if it exists
            try:
                elements = await page.query_selector_all('button, [role="button"], a')
                for el in elements:
                    el_text = await el.inner_text()
                    if el_text and 'Book details' in el_text:
                        await el.click()
                        print(f"  [Row {row_idx}] Clicked 'Book details' dropdown to reveal publisher!")
                        await asyncio.sleep(1.5)
                        break
            except Exception:
                pass
            
            # 4. Extract Publisher
            agency = ""
            pub_info = await page.query_selector('[data-testid="publicationInfo"]')
            if pub_info:
                pub_text = await pub_info.inner_text()
                if " by " in pub_text:
                    agency = pub_text.split(" by ")[-1].strip()
            
            # Fallback for Classic HTML layout (usually inside #details .row)
            if not agency:
                try:
                    # In classic HTML, it's usually inside <div id="details"><div class="row">
                    rows = await page.query_selector_all('#details .row, .FeaturedDetails, .infoBoxRowItem')
                    for r in rows:
                        txt = await r.inner_text()
                        if "Published" in txt and " by " in txt:
                            agency = txt.split(" by ")[-1].strip()
                            # Clean up any trailing text that might be captured (e.g. newlines)
                            agency = agency.split('\n')[0].strip()
                            break
                except Exception:
                    pass

            # Final Regex Fallback across the entire body text
            if not agency:
                body_text = await page.evaluate("document.body.innerText")
                # Look for "Published [Date] by [Publisher]" handling newlines
                match = re.search(r'Published.{1,50}?by\s+([^\n]+)', body_text, re.IGNORECASE | re.DOTALL)
                if match:
                    agency = match.group(1).strip()
                    
            if agency:
                print(f"  [Row {row_idx}] Extracted Publisher: {agency}")
                async with excel_lock:
                    ws.cell(row=row_idx, column=col_map['Agency / Publisher']).value = agency
                    wb.save(INPUT_FILE)
            else:
                print(f"  [Row {row_idx}] Could not find publisher details on page.")
                
        except BlockedError as e:
            print(f"  [Row {row_idx}] BLOCKED by Goodreads! {e}")
        except Exception as e:
            print(f"  [Row {row_idx}] Failed to process: {e}")
        finally:
            await page.close()

async def main():
    print(f"Loading {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE)
    if 'Sheet1' not in wb.sheetnames:
        print("Sheet1 not found!")
        return
    ws = wb['Sheet1']
    
    header_row = 1
    col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            col_map[str(val).strip()] = col_idx
            
    required_cols = ['Goodreads Series URL', 'Agency / Publisher']
    for req in required_cols:
        if req not in col_map:
            print(f"Required column '{req}' not found in headers!")
            return
            
    semaphore = asyncio.Semaphore(2)  # Reduced to 2 to prevent timeout bans
    excel_lock = asyncio.Lock()
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        print("Starting Login Phase...")
        login_page = await context.new_page()
        scraper = GoodreadsScraper(headless=False)
        await scraper.login_to_goodreads(login_page)
        await login_page.close()
        print("Login complete. Starting concurrent scrape...")
        
        tasks = []
        for row_idx in range(53, 84): # Next 30 rows (up to 83)
            tasks.append(process_row(row_idx, ws, col_map, context, semaphore, excel_lock, wb))
            
        if tasks:
            await asyncio.gather(*tasks)
        else:
            print("No tasks found in the specified range.")
            
        await browser.close()
    
    print(f"Scraping run complete.")

if __name__ == "__main__":
    asyncio.run(main())
