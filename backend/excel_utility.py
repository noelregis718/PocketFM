import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def save_to_excel(data, filename="scraped_data_keywords.xlsx"):
    # Define the exact order of columns requested by the user
    columns = [
        "Sub_Genre", "Price_Tier", "Amazon URL", "Book Title", 
        "Book Number in Series", "Series Name", "Author Name", 
        "Amazon Stars", "Amazon Ratings", "Number of Books in Series", 
        "Genre", "Publisher", "Publication Date", "Print Length / Pages", 
        "Best Sellers Rank", "Licensing Status", "Part of a Series?", 
        "Part_of_Series", "GoodReads_Series_URL", "Num_Primary_Books", 
        "Total_Pages_Primary_Books", "Book1_Rating", "Book1_Num_Ratings", 
        "Logline", "One_Sentence_Logline", "Romantasy_Subgenre", 
        "Author_Email", "Agent_Email", "Facebook", "Twitter", 
        "Instagram", "Website", "Other_Contact"
    ]

    # --- INDUSTRIAL STEPPER: APPEND MODE ---
    df_new = pd.DataFrame(data)
    
    # Ensure all required columns are present in df_new, filled with N/A if missing
    for col in columns:
        if col not in df_new.columns:
            df_new[col] = "N/A"
            
    df_new = df_new.reindex(columns=columns)
    
    # Check if file exists to determine append or fresh start
    if os.path.exists(filename):
        try:
            df_existing = pd.read_excel(filename)
            # Ensure columns match
            df_existing = df_existing.reindex(columns=columns)
            # Combine and deduplicate based on Amazon URL
            df = pd.concat([df_existing, df_new], ignore_index=True)
            df = df.drop_duplicates(subset=['Amazon URL'], keep='last')
        except Exception as e:
            print(f"  [Warning] Could not append to {filename}: {e}. Starting fresh.")
            df = df_new
    else:
        df = df_new

    def _write(path):
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Amazon Scraped Books')

            worksheet = writer.sheets['Amazon Scraped Books']
            
            # Freeze the top row
            worksheet.freeze_panes = "A2"

            # Styles
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )

            # Per-column width rules for the new 33-column schema
            col_widths = {
                'Sub_Genre': 15, 'Price_Tier': 20, 'Amazon URL': 40, 'Book Title': 40,
                'Book Number in Series': 12, 'Series Name': 25, 'Author Name': 25,
                'Amazon Stars': 10, 'Amazon Ratings': 15, 'Number of Books in Series': 15,
                'Genre': 15, 'Publisher': 25, 'Publication Date': 18, 'Print Length / Pages': 15,
                'Best Sellers Rank': 40, 'Licensing Status': 15, 'Part of a Series?': 12,
                'Part_of_Series': 15, 'GoodReads_Series_URL': 25, 'Num_Primary_Books': 12,
                'Total_Pages_Primary_Books': 15, 'Book1_Rating': 12, 'Book1_Num_Ratings': 12,
                'Logline': 50, 'One_Sentence_Logline': 45, 'Romantasy_Subgenre': 20,
                'Author_Email': 25, 'Agent_Email': 25, 'Facebook': 15, 'Twitter': 15,
                'Instagram': 15, 'Website': 20, 'Other_Contact': 20
            }

            for idx, col in enumerate(df.columns):
                col_letter = get_column_letter(idx + 1)
                width = col_widths.get(col, 20)
                worksheet.column_dimensions[col_letter].width = width

                for cell in worksheet[col_letter]:
                    # Default alignment
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    # Add border to every cell
                    cell.border = thin_border
                    
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            # Dynamic row height: auto-fit based on content length
            for row_num in range(2, worksheet.max_row + 1):
                max_lines = 1
                for col_idx, col_name in enumerate(df.columns):
                    col_letter = get_column_letter(col_idx + 1)
                    cell = worksheet[f"{col_letter}{row_num}"]
                    val = str(cell.value) if cell.value is not None else ""
                    col_width = col_widths.get(col_name, 20)

                    # Count explicit newlines
                    newline_count = val.count('\n') + 1

                    # Estimate wrapped lines
                    chars_per_line = max(int(col_width * 1.1), 10)
                    for line in val.split('\n'):
                        wrapped = max(1, -(-len(line) // chars_per_line))
                        newline_count += wrapped - 1

                    max_lines = max(max_lines, newline_count)

                worksheet.row_dimensions[row_num].height = max(min(max_lines * 15, 300), 20)

    # Primary attempt: write to the requested filename
    try:
        _write(filename)
        print(f"Excel saved: {filename}")
        return os.path.abspath(filename)
    except PermissionError:
        # File is open in Excel — use a timestamped fallback filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(filename)
        fallback = f"{base}_{timestamp}{ext}"
        print(f"WARNING: '{filename}' is open in Excel. Saving to '{fallback}' instead.")
        _write(fallback)
        return os.path.abspath(fallback)

