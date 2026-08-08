import openpyxl

file_path = r'e:\Internship\PocketFM\Merged_Romance_Keywords.xlsx'
print(f"Loading {file_path}...")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Target columns
col_I = 9
col_J = 10
col_K = 11
col_L = 12

# Source columns
col_X = 24
col_Y = 25
col_Z = 26
col_AA = 27

print("Merging data...")
for row in range(2, ws.max_row + 1):
    # Merge X into I
    val_I = ws.cell(row=row, column=col_I).value
    val_X = ws.cell(row=row, column=col_X).value
    if (val_I is None or val_I == "") and val_X is not None:
        ws.cell(row=row, column=col_I).value = val_X

    # Merge Y into J
    val_J = ws.cell(row=row, column=col_J).value
    val_Y = ws.cell(row=row, column=col_Y).value
    if (val_J is None or val_J == "") and val_Y is not None:
        ws.cell(row=row, column=col_J).value = val_Y

    # Merge Z into K
    val_K = ws.cell(row=row, column=col_K).value
    val_Z = ws.cell(row=row, column=col_Z).value
    if (val_K is None or val_K == "") and val_Z is not None:
        ws.cell(row=row, column=col_K).value = val_Z

    # Merge AA into L
    val_L = ws.cell(row=row, column=col_L).value
    val_AA = ws.cell(row=row, column=col_AA).value
    if (val_L is None or val_L == "") and val_AA is not None:
        ws.cell(row=row, column=col_L).value = val_AA

print("Saving workbook...")
wb.save(file_path)
print("Data merged successfully (columns not deleted).")
