import openpyxl

file_path = r'e:\Internship\PocketFM\Merged_Romance_Keywords.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

headers = []
for col in range(1, ws.max_column + 1):
    headers.append(ws.cell(row=1, column=col).value)

print("Current Headers:")
for i, h in enumerate(headers):
    print(f"Col {i+1}: {h}")
