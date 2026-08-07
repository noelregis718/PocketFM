import openpyxl

file_path = r'e:\Internship\PocketFM\Merged_Romance_Keywords.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Define new columns
new_columns = ["Goodreads Rating", "Goodreads No. of Ratings", "Goodreads Link"]

# Get the max column
max_col = ws.max_column

# Get the header formatting from the first column to apply it to new columns
header_cell_reference = ws.cell(row=1, column=1)

for i, col_name in enumerate(new_columns):
    new_col_idx = max_col + i + 1
    new_cell = ws.cell(row=1, column=new_col_idx)
    new_cell.value = col_name
    
    # Copy styling
    if header_cell_reference.has_style:
        new_cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF", name='Calibri', size=11)
        new_cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        new_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        new_cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin', color='BFBFBF'), 
            right=openpyxl.styles.Side(style='thin', color='BFBFBF'), 
            top=openpyxl.styles.Side(style='thin', color='BFBFBF'), 
            bottom=openpyxl.styles.Side(style='thin', color='BFBFBF')
        )
        
    # set width
    col_letter = openpyxl.utils.get_column_letter(new_col_idx)
    ws.column_dimensions[col_letter].width = 18

print("Adding borders to the rest of the new columns...")
# Add borders to all rows for the new columns
for row in range(2, ws.max_row + 1):
    for i in range(len(new_columns)):
        col_idx = max_col + i + 1
        cell = ws.cell(row=row, column=col_idx)
        cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin', color='BFBFBF'), 
            right=openpyxl.styles.Side(style='thin', color='BFBFBF'), 
            top=openpyxl.styles.Side(style='thin', color='BFBFBF'), 
            bottom=openpyxl.styles.Side(style='thin', color='BFBFBF')
        )

wb.save(file_path)
print("New Goodreads columns added successfully.")
