import openpyxl

file_path = 'Amazon A-Z Crawl List.xlsx'
print('Loading workbook...')
wb = openpyxl.load_workbook(file_path)
keep_sheet = 'Romance + Fantasy - Cut 1 (Uniq'
print('Current sheets:', wb.sheetnames)

sheets_to_remove = [s for s in wb.sheetnames if s != keep_sheet]
for s in sheets_to_remove:
    del wb[s]

wb.save(file_path)
print('Successfully saved. Remaining sheets:', wb.sheetnames)
