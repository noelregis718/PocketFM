import openpyxl

file_path = r'e:\Internship\PocketFM\Merged_Romance_Keywords.xlsx'
print(f"Loading {file_path}...")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Column indices (1-based)
col_I = 9
col_J = 10
col_K = 11
col_L = 12

col_V = 22
col_W = 23
col_Y = 25
col_Z = 26

print("Merging data...")
for row in range(2, ws.max_row + 1):
    # 1. Merge V into I
    val_I = ws.cell(row=row, column=col_I).value
    val_V = ws.cell(row=row, column=col_V).value
    if (val_I is None or val_I == "") and val_V is not None:
        ws.cell(row=row, column=col_I).value = val_V

    # 2. Merge W into J
    val_J = ws.cell(row=row, column=col_J).value
    val_W = ws.cell(row=row, column=col_W).value
    if (val_J is None or val_J == "") and val_W is not None:
        ws.cell(row=row, column=col_J).value = val_W

    # 3. Merge Y into K
    val_K = ws.cell(row=row, column=col_K).value
    val_Y = ws.cell(row=row, column=col_Y).value
    if (val_K is None or val_K == "") and val_Y is not None:
        ws.cell(row=row, column=col_K).value = val_Y

    # 4. Merge Z into L
    val_L = ws.cell(row=row, column=col_L).value
    val_Z = ws.cell(row=row, column=col_Z).value
    if (val_L is None or val_L == "") and val_Z is not None:
        ws.cell(row=row, column=col_L).value = val_Z

# Columns to delete (in descending order to avoid index shifting issues)
# 31: AE (Licensing Status)
# 30: AD (Total_Pages_Primary_Books)
# 29: AC (Num_Primary_Books)
# 27: AA (Number of reviews)
# 26: Z (Amazon Ratings)
# 25: Y (Amazon Stars)
# 23: W (Print Length / Pages)
# 22: V (Number of Books in Series)
cols_to_delete = [31, 30, 29, 27, 26, 25, 23, 22]

print("Deleting columns...")
for col_idx in cols_to_delete:
    print(f"Deleting column {col_idx} ({ws.cell(row=1, column=col_idx).value})...")
    ws.delete_cols(col_idx)

print("Saving workbook...")
wb.save(file_path)
print("Data merged and columns deleted successfully.")
