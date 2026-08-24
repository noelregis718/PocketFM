import asyncio
import json
import os
import pandas as pd
import re
from scraper import AmazonScraper, clean_text, normalize_title_for_search
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font

MASTER_STATE_FILE = r"E:\Internship\PocketFM\backend\master_mission_state_amazon_litrpg.json"
MISSION_KEY = "litrpg_romance"
OUTPUT_FILE = r"E:\Internship\PocketFM\Combined List of Titles.xlsx"

BATCH_SIZE = 500
MAX_TABS = 8
MAX_TOTAL_LIMIT = 100000
SEARCH_URL = "https://www.amazon.com/s?k=LitRPG+Romance&i=stripbooks&crid=10JFR8ZJP0WSW&sprefix=%2Cstripbooks%2C359&ref=nb_sb_noss_2"
SUB_GENRE = "Korean Romance Fantasy / Isekai"
KEYWORD = "LitRPG Romance"
MAX_PAGES = 110

def ensure_excel(output_file):
    if not os.path.exists(output_file):
        columns = [
            "Book Title", "Author Name", "Series Name", "Sub-Genre", "Logline", 
            "Amazon URL", "Keyword", "Amazon Num_Primary_Books_in_Series", 
            "Amazon Total_Page_Count_of_Primary_Books", "Amazon Book1_Rating", 
            "Amazon Book1_Num_Ratings", "Publisher", "Publication Date", "Best Sellers Rank", 
            "GoodReads_Series_URL", "Goodreads Primary Books Number", "Goodreads Primary Books Page Count", 
            "Goodreads Rating Book 1", "Goodreads No. of Ratings Book 1", "Goodreads Link", 
            "Genre Tags", "Romantasy Checker", "Synopsis"
        ]
        df = pd.DataFrame(columns=columns)
        df.to_excel(output_file, index=False)
        print(f"Created fresh Excel sheet: {output_file}")

def load_state():
    if os.path.exists(MASTER_STATE_FILE):
        with open(MASTER_STATE_FILE, 'r') as f:
            try:
                master = json.load(f)
                if MISSION_KEY in master:
                    state = master[MISSION_KEY]
                    state['last_page_scanned'] = int(state.get('last_page_scanned', 0))
                    state['total_processed_global'] = int(state.get('total_processed_global', 0))
                    state['last_asin'] = state.get('last_asin', None)
                    state['last_book_title'] = state.get('last_book_title', None)
                    return state
            except json.JSONDecodeError:
                pass
    return {"last_page_scanned": 0, "total_processed_global": 0, "last_asin": None, "last_book_title": None}

def save_state(state):
    master = {}
    if os.path.exists(MASTER_STATE_FILE):
        try:
            with open(MASTER_STATE_FILE, 'r') as f:
                master = json.load(f)
        except json.JSONDecodeError:
            pass
    master[MISSION_KEY] = state
    with open(MASTER_STATE_FILE, 'w') as f:
        json.dump(master, f, indent=4)

def extract_series_custom(title):
    if not title: return ""
    # "New Series Book 1 - Hear me Out" -> "New Series"
    match_dash = re.search(r'^(.*?) Book \d+ -', title, re.IGNORECASE)
    if match_dash: return match_dash.group(1).strip()
    
    # Inverted commas: "Series Name" or 'Series Name'
    match_quotes = re.search(r'["\'](.*?)["\']', title)
    if match_quotes: return match_quotes.group(1).strip()
    
    # Parentheses: (Series Name Book 1)
    match_paren = re.search(r'\((.*?)(?:\s+#?\d+|\s+Book\s+\d+)?\)', title, re.IGNORECASE)
    if match_paren:
        name = match_paren.group(1).strip()
        name = re.sub(r'[\s#]+$', '', name)
        if len(name) > 2: return name
    return ""

def apply_styling(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
            
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 15
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
        wb.save(excel_path)
        print("Excel styling and text wrapping applied successfully.")
    except Exception as e:
        print(f"Failed to apply styling: {e}")

async def check_captcha(page):
    while True:
        try:
            if await page.query_selector('form[action="/errors/validateCaptcha"], input#captchacharacters'):
                print("🚨 CAPTCHA DETECTED! Please solve it in the browser window...", flush=True)
                await asyncio.sleep(5)
            else:
                break
        except:
            break

async def check_continue_shopping(page):
    try:
        btn = await page.query_selector("text='Continue shopping'")
        if not btn:
            btn = await page.query_selector("a:has-text('Continue shopping')")
        if btn:
            print("Found 'Continue shopping' button. Clicking it...", flush=True)
            await btn.click()
            await asyncio.sleep(2)
    except Exception as e:
        pass

async def process_book(context, book_data):
    amazon = AmazonScraper()
    url = book_data.get("Amazon URL")
    discovery_title = book_data.get("Book Title", "N/A")
    asin = book_data.get("asin", "N/A")

    amz_details = await amazon.scrape_product_details_tab(context, url)
    
    price_raw = amz_details.get("Price", "N/A")
    if "INR" in price_raw or "₹" in price_raw or "\u20b9" in price_raw or "Rs" in price_raw:
        print(f"    [Heartbeat] Non-USD detected. Location sync...")
        try:
            temp_page = await context.new_page()
            await temp_page.goto("https://www.amazon.com", wait_until="domcontentloaded", timeout=45000)
            await check_continue_shopping(temp_page)
            await amazon.set_amazon_location(temp_page, "90016")
            await asyncio.sleep(2)
            await temp_page.close()
            amz_details = await amazon.scrape_product_details_tab(context, url)
        except Exception as e:
            pass

    actual_title = amz_details.get("Book Title") if (amz_details.get("Book Title") and amz_details.get("Book Title") != "N/A") else discovery_title
    author_name = amz_details.get("Author Name", "N/A")
    print(f"  [Task] Processed: {actual_title[:40]}... (ASIN: {asin})", flush=True)

    desc = amz_details.get("Description", "N/A")
    publisher = amz_details.get("Publisher", "N/A")
    pub_date = amz_details.get("Publication Date", "N/A")

    series_name = amz_details.get("Series Name", "")
    if not series_name or series_name == "N/A":
        series_name = extract_series_custom(actual_title)
        
    num_primary_books = amz_details.get("Total Books", "")
    total_pages_primary = amz_details.get("Pages", "")

    result = {
        "Book Title": actual_title,
        "Author Name": author_name,
        "Series Name": series_name,
        "Sub-Genre": SUB_GENRE,
        "Logline": desc[:1500] if desc != "N/A" else "",
        "Amazon URL": url,
        "Keyword": KEYWORD,
        "Amazon Num_Primary_Books_in_Series": num_primary_books,
        "Amazon Total_Page_Count_of_Primary_Books": total_pages_primary,
        "Amazon Book1_Rating": amz_details.get("Rating", ""),
        "Amazon Book1_Num_Ratings": amz_details.get("Number of Reviews", ""),
        "Publisher": publisher,
        "Publication Date": pub_date,
        "Best Sellers Rank": amz_details.get("Inner Rank", ""),
        "GoodReads_Series_URL": "",
        "Goodreads Primary Books Number": "",
        "Goodreads Primary Books Page Count": "",
        "Goodreads Rating Book 1": "",
        "Goodreads No. of Ratings Book 1": "",
        "Goodreads Link": "",
        "Genre Tags": "",
        "Romantasy Checker": "",
        "Synopsis": ""
    }

    for k, v in result.items():
        if v == "N/A":
            result[k] = ""
            
    return result

async def run_scraper():
    ensure_excel(OUTPUT_FILE)
    state = load_state()
    
    if state['last_page_scanned'] > 0:
        print(f"\n[Status Loaded] Resuming from Page {state['last_page_scanned']}...")
        if state['last_book_title']:
            print(f"[Status Loaded] Last processed book: '{state['last_book_title']}' (ASIN: {state['last_asin']})\n")

    global_seen_asins = set()
    global_seen_titles = set()
    
    if os.path.exists(OUTPUT_FILE):
        try:
            existing_df = pd.read_excel(OUTPUT_FILE)
            if 'Amazon URL' in existing_df.columns:
                for url in existing_df['Amazon URL'].dropna():
                    match = re.search(r'/(?:dp|product|gp/product)/([A-Z0-9]{10})', str(url))
                    if match: global_seen_asins.add(match.group(1))
            if 'Book Title' in existing_df.columns:
                for t in existing_df['Book Title'].dropna():
                    global_seen_titles.add(normalize_title_for_search(str(t)))
            print(f"Loaded {len(global_seen_asins)} existing ASINs to prevent duplicates.", flush=True)
        except Exception as e:
            print(f"Could not load existing Excel: {e}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
        )

        while True:
            if state['total_processed_global'] >= MAX_TOTAL_LIMIT:
                print(f"Reached test limit of {MAX_TOTAL_LIMIT} books. Stopping.", flush=True)
                break
            
            page = await context.new_page()
            search_url = SEARCH_URL
            try:
                print(f"Navigating to Base Amazon Search...", flush=True)
                await page.goto(search_url, wait_until="domcontentloaded", timeout=60000)
                await check_continue_shopping(page)
                amazon_scraper = AmazonScraper()
                await amazon_scraper.set_amazon_location(page, "90016")
                await check_captcha(page)

                target_page = int(state['last_page_scanned']) + 1
                if target_page > 1:
                    if "&page=" in search_url:
                        search_url = re.sub(r'&page=\d+', f'&page={target_page}', search_url)
                    else:
                        search_url += f"&page={target_page}"
                    if "&ref=sr_pg_" in search_url:
                        search_url = re.sub(r'&ref=sr_pg_\d+', f'&ref=sr_pg_{target_page}', search_url)
                    else:
                        search_url += f"&ref=sr_pg_{target_page}"
                    print(f"Navigating directly to Target Amazon Search (Page {target_page})...", flush=True)
                    await page.goto(search_url, wait_until="load", timeout=60000)
                    await check_continue_shopping(page)
                    await check_captcha(page)

                all_discovery_links = []
                page_count = target_page
                
                if page_count >= MAX_PAGES:
                    print(f"Reached maximum page limit of {MAX_PAGES}. Stopping scraper.", flush=True)
                    break
                
                print(f"Scanning Page {page_count} for new titles...", flush=True)
                while len(all_discovery_links) < BATCH_SIZE:
                    for i in range(5):
                        await page.evaluate("window.scrollBy(0, 1500)")
                        await asyncio.sleep(1.5)

                    items = await page.query_selector_all("[data-asin]")
                    if not items:
                        print("No items found. End of results or CAPTCHA blocking.", flush=True)
                        break
                        
                    for item in items:
                        asin = await item.get_attribute('data-asin')
                        if not asin or len(asin) < 5 or asin in global_seen_asins: continue
                        if any(x.get("asin") == asin for x in all_discovery_links): continue

                        title = "N/A"
                        for t_sel in ["h2 a span", ".a-size-medium", ".a-size-base-plus"]:
                            try:
                                t_el = await item.query_selector(t_sel)
                                if t_el:
                                    title = clean_text(await t_el.inner_text())
                                    if title and title != "N/A": break
                            except: continue

                        clean_title = normalize_title_for_search(title)
                        if clean_title in global_seen_titles: continue

                        href = None
                        for l_sel in ["h2 a", "a.a-link-normal[href*='/dp/']"]:
                            try:
                                l_el = await item.query_selector(l_sel)
                                if l_el:
                                    href = await l_el.evaluate("el => el.href")
                                    if href and "/dp/" in href: break
                            except: continue

                        if href and title != "N/A":
                            all_discovery_links.append({"asin": asin, "Amazon URL": href, "Book Title": title})
                        if len(all_discovery_links) >= BATCH_SIZE: break

                    print(f" -> Found {len(all_discovery_links)}/{BATCH_SIZE} links...", flush=True)
                    if len(all_discovery_links) < BATCH_SIZE:
                        page_count += 1
                        
                        if page_count > MAX_PAGES:
                            print(f"Reached maximum page limit of {MAX_PAGES} while scanning. Stopping discovery phase.", flush=True)
                            break
                            
                        if "&page=" in search_url:
                            search_url = re.sub(r'&page=\d+', f'&page={page_count}', search_url)
                        else:
                            search_url += f"&page={page_count}"
                        
                        try:
                            print(f"Flipping to Page {page_count}...", flush=True)
                            await page.goto(search_url, wait_until="load", timeout=60000)
                            await check_continue_shopping(page)
                            await check_captcha(page)
                            state['last_page_scanned'] = page_count - 1
                        except Exception as e:
                            print(f"Failed to navigate to next page: {e}", flush=True)
                            break

                final_rows = []
                for i in range(0, len(all_discovery_links), MAX_TABS):
                    batch = all_discovery_links[i : i + MAX_TABS]
                    print(f"Processing Batch {i//MAX_TABS + 1} ({len(batch)} items)...", flush=True)
                    tasks = [asyncio.wait_for(process_book(context, book), timeout=120) for book in batch]
                    results = await asyncio.gather(*tasks, return_exceptions=True)
                    valid = [res for res in results if isinstance(res, dict)]
                    final_rows.extend(valid)
                
                # Append to excel
                if final_rows:
                    new_df = pd.DataFrame(final_rows)
                    if os.path.exists(OUTPUT_FILE):
                        old_df = pd.read_excel(OUTPUT_FILE)
                        combined = pd.concat([old_df, new_df], ignore_index=True)
                    else:
                        combined = new_df
                    
                    if 'Amazon URL' in combined.columns:
                        combined = combined.drop_duplicates(subset=['Amazon URL'], keep='first')
                    if 'Book Title' in combined.columns:
                        combined = combined.drop_duplicates(subset=['Book Title'], keep='first')
                        
                    combined.to_excel(OUTPUT_FILE, index=False)
                    apply_styling(OUTPUT_FILE)

                    global_seen_asins.update([r.get('asin') for r in all_discovery_links if r.get('asin')])
                    state['last_page_scanned'] = page_count
                    state['total_processed_global'] += len(final_rows)
                    
                    last_book_processed = final_rows[-1]
                    state['last_book_title'] = last_book_processed.get('Book Title', '')
                    
                    last_asin = None
                    for dl in all_discovery_links:
                        if dl.get('Amazon URL') == last_book_processed.get('Amazon URL'):
                            last_asin = dl.get('asin')
                    state['last_asin'] = last_asin
                    
                    save_state(state)
                    print(f"Batch saved! Total Processed: {state['total_processed_global']}", flush=True)
                else:
                    print("No books were extracted in this iteration.", flush=True)
                    if all_discovery_links:
                        print("Possible block. Pausing for a moment.", flush=True)
                        await asyncio.sleep(10)
                    else:
                        break
                    
            except Exception as e:
                print(f"Error in batch loop: {e}", flush=True)
                await asyncio.sleep(5)
            
            await page.close()

        await browser.close()
        print("Scraping completed!", flush=True)

if __name__ == "__main__":
    asyncio.run(run_scraper())
