import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

file_path = r'e:\Internship\PocketFM\Merged_Romance_Keywords.xlsx'
print(f"Loading {file_path}...")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Define styles
header_font = Font(bold=True, color="FFFFFF", name='Calibri', size=11)
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
alignment_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin_border = Border(left=Side(style='thin', color='BFBFBF'), 
                     right=Side(style='thin', color='BFBFBF'), 
                     top=Side(style='thin', color='BFBFBF'), 
                     bottom=Side(style='thin', color='BFBFBF'))
regular_font = Font(name='Calibri', size=11)
hyperlink_font = Font(color="0563C1", underline="single", name='Calibri', size=11)

print("Formatting header...")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment_center
    cell.border = thin_border

print("Formatting columns...")
for col in range(1, ws.max_column + 1):
    letter = get_column_letter(col)
    cell_value = ws.cell(row=1, column=col).value
    header_str = str(cell_value).lower() if cell_value else ""
    
    if "synopsis" in header_str or "description" in header_str or "summary" in header_str:
        ws.column_dimensions[letter].width = 60
    elif "url" in header_str or "link" in header_str:
        ws.column_dimensions[letter].width = 40
    elif "name" in header_str or "title" in header_str:
        ws.column_dimensions[letter].width = 30
    else:
        ws.column_dimensions[letter].width = 20

print("Applying data styles...")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = alignment_left_top
        cell.border = thin_border
        
        # Identify links
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('http'):
            cell.font = hyperlink_font
            cell.hyperlink = cell.value
        else:
            cell.font = regular_font

# Setting fixed row height as in style_wrap.py
print("Setting row heights...")
for row_idx in range(2, ws.max_row + 1):
    ws.row_dimensions[row_idx].height = 15

ws.freeze_panes = 'A2'

print("Saving workbook...")
wb.save(file_path)
print("Styling applied successfully!")
