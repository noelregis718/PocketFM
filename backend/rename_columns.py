import openpyxl

file_path = r'e:\Internship\PocketFM\Merged_Romance_Keywords.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

col_K_val = ws['K1'].value
col_L_val = ws['L1'].value

print(f"Original Column K: {col_K_val}")
print(f"Original Column L: {col_L_val}")

def prepend_amazon(val):
    if val and isinstance(val, str):
        if "amazon" not in val.lower():
            return f"Amazon {val}"
    return val

ws['K1'].value = prepend_amazon(col_K_val)
ws['L1'].value = prepend_amazon(col_L_val)

print(f"New Column K: {ws['K1'].value}")
print(f"New Column L: {ws['L1'].value}")

wb.save(file_path)
print("Headers updated successfully.")
