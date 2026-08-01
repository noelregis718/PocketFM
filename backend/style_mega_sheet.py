import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

filename = 'Mega_Sheet.xlsx'
print(f"Loading {filename}...")
wb = openpyxl.load_workbook(filename)
ws = wb.active

# Define styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

print("Applying header styles...")
# Apply header styles
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")

print("Applying borders to all data and setting thin row heights...")
# Set thin rows (row height = 12) and apply borders to all cells
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border

# Explicitly set row heights to 12.5 (a bit thinner than standard 15) for all rows
for row_idx in range(1, ws.max_row + 1):
    ws.row_dimensions[row_idx].height = 12.5

print("Adjusting column widths and adding filters...")
# Adjust column widths based on header length to ensure readability
for col in range(1, ws.max_column + 1):
    column_letter = get_column_letter(col)
    header_val = str(ws.cell(row=1, column=col).value)
    ws.column_dimensions[column_letter].width = max(len(header_val) + 2, 12)

# Freeze top row and add auto-filters
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

print("Saving workbook...")
wb.save(filename)
print("Styling applied successfully!")
