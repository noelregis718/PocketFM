import asyncio
import os
import re
import urllib.parse
import json
import random
import openpyxl
from playwright.async_api import async_playwright

EXCEL_FILE = r"E:\Internship\PocketFM\Merged_Romance_Keywords.xlsx"
SHEET_NAME = "Sheet1" # Default active sheet

async def process_row(row, context, sem, ws, wb, headers, lock):
    async with sem:
        await asyncio.sleep(random.uniform(1.0, 3.0)) # Jitter
        
        title = ws.cell(row=row, column=headers.get('Book Title', 0)).value
        series = ws.cell(row=row, column=headers.get('Series Name', 0)).value
        author = ws.cell(row=row, column=headers.get('Author Name', 0)).value
        
        book_name = str(title if title else (series if series else "")).strip()
        if ":" in book_name:
            book_name = book_name.split(":")[0].strip()
        author_name = str(author if author else "").strip()
        
        if not book_name or book_name.lower() == 'nan':
            return
            
        print(f"\n--- [Row {row}] Processing '{book_name}' by '{author_name}' ---")
            
        search_query = f"{book_name} {author_name}".strip()
        search_url = f"https://www.goodreads.com/search?q={urllib.parse.quote_plus(search_query)}"
        
        print(f"[Row {row}] [SEARCH] Searching: {search_query}")
        
        b_url_initial = None
        search_page = await context.new_page()
        try:
            await search_page.goto(search_url, wait_until="domcontentloaded", timeout=45000)
            await asyncio.sleep(2)
            first_book_el = await search_page.query_selector('a[href*="/book/show/"]')
            if not first_book_el:
                print(f"[Row {row}] [SEARCH] No book links found in search.")
                return
            b_url_initial = await first_book_el.get_attribute("href")
            if not b_url_initial.startswith('http'): 
                b_url_initial = "https://www.goodreads.com" + b_url_initial
        except Exception as e:
            print(f"[Row {row}] [SEARCH] Error: {e}")
            raise
        finally:
            await search_page.close()
            
        if not b_url_initial:
            return
            
        # 2. Book Page Extraction
        print(f"[Row {row}] [BOOK] Extracting: {b_url_initial}")
        book_page = await context.new_page()
        
        extracted_data = {
            "Goodreads Link": b_url_initial,
            "Goodreads Rating Book 1": None,
            "Goodreads No. of Ratings Book 1": None,
            "Publisher": None,
            "Publication Date": None,
            "Synopsis": None,
            "Genre Tags": None,
            "GoodReads_Series_URL": None
        }
        
        try:
            await book_page.goto(b_url_initial, wait_until="domcontentloaded", timeout=45000)
            await asyncio.sleep(2)
            
            # Genres
            try:
                genre_btns = await book_page.query_selector_all('[data-testid="genresList"] button, [data-testid="genresList"] [role="button"]')
                for btn in genre_btns:
                    if "...more" in (await btn.inner_text()).lower():
                        await btn.click(force=True)
                        await asyncio.sleep(0.5)
                        break
            except: pass
            
            genres = []
            genre_els = await book_page.query_selector_all('[data-testid="genresList"] .Button__labelItem, .BookPageMetadataSection__genre a')
            for gel in genre_els:
                txt = (await gel.inner_text()).strip()
                if txt and txt not in genres: genres.append(txt)
            if genres: extracted_data["Genre Tags"] = ", ".join(genres)
            
            # Synopsis/Logline
            desc_el = await book_page.query_selector('[data-testid="description"] .Formatted, .readable')
            if desc_el: extracted_data["Synopsis"] = (await desc_el.inner_text()).strip()
            
            # Ratings
            rating_el = await book_page.query_selector('div.RatingStatistics__rating')
            if rating_el:
                try: extracted_data["Goodreads Rating Book 1"] = float((await rating_el.inner_text()).strip())
                except: pass
            
            count_el = await book_page.query_selector('[data-testid="ratingsCount"]')
            if count_el:
                try:
                    count_txt = (await count_el.inner_text()).replace(',', '').split()[0]
                    extracted_data["Goodreads No. of Ratings Book 1"] = int(count_txt)
                except: pass
                
            # JSON-LD for Publisher and Date
            ld_el = await book_page.query_selector('script[type="application/ld+json"]')
            if ld_el:
                try:
                    data = json.loads(await ld_el.inner_text())
                    if isinstance(data, list): data = data[0]
                    if 'publisher' in data:
                        extracted_data["Publisher"] = data['publisher']
                    if 'datePublished' in data:
                        extracted_data["Publication Date"] = data['datePublished']
                except: pass
                
            # Fallback for Publisher/Date via UI
            if not extracted_data["Publisher"] or not extracted_data["Publication Date"]:
                pub_el = await book_page.query_selector('[data-testid="publicationInfo"]')
                if pub_el:
                    pub_text = await pub_el.inner_text() # e.g. "Published March 4, 2014 by Tor Books"
                    if "by" in pub_text and not extracted_data["Publisher"]:
                        raw_pub = pub_text.split("by")[-1].strip()
                        # Clean up Publisher string: take only the first line to drop garbage UI text
                        extracted_data["Publisher"] = raw_pub.split("\n")[0].strip()
                    # Try to extract date
                    date_match = re.search(r'Published\s+(.*?)\s+by', pub_text, re.IGNORECASE)
                    if date_match and not extracted_data["Publication Date"]:
                        extracted_data["Publication Date"] = date_match.group(1).strip()
            
            # Series Link
            series_el = await book_page.query_selector('a[href*="/series/"]')
            if series_el:
                series_link = await series_el.get_attribute("href")
                if not series_link.startswith('http'): series_link = "https://www.goodreads.com" + series_link
                extracted_data["GoodReads_Series_URL"] = series_link
                
        except Exception as e:
            print(f"[Row {row}] [BOOK] Error: {e}")
            raise
        finally:
            await book_page.close()
            
        # 3. Series Page Logic
        extracted_data["Goodreads Primary Books Number"] = 0
        extracted_data["Goodreads Primary Books Page Count"] = 0
        
        if extracted_data["GoodReads_Series_URL"]:
            print(f"[Row {row}] [SERIES] Opening Series Page: {extracted_data['GoodReads_Series_URL']}")
            series_page = await context.new_page()
            try:
                await series_page.goto(extracted_data["GoodReads_Series_URL"], wait_until="domcontentloaded", timeout=45000)
                await asyncio.sleep(2)
                
                book_items = await series_page.query_selector_all('.listWithDividers__item')
                books_to_scrape = []
                
                for item in book_items:
                    title_el = await item.query_selector('a.bookTitle, span[itemprop="name"]')
                    if not title_el: title_el = await item.query_selector('a[href*="/book/show/"]')
                    if not title_el: continue
                    
                    item_text = await title_el.inner_text()
                    h3_el = await item.query_selector('h3')
                    h3_text = await h3_el.inner_text() if h3_el else ""
                    
                    book_num = None
                    match = re.search(r'\((?:.+?)(?:,\s*#|,\s*Book\s+|\s+Book\s+|\s+#)([0-9a-zA-Z\.\-]+)\)', item_text, re.IGNORECASE)
                    if match:
                        if match.group(1).isdigit(): book_num = match.group(1)
                    if not book_num and h3_text:
                        match2 = re.search(r'Book\s+(\d+)$', h3_text.strip(), re.IGNORECASE)
                        if match2: book_num = match2.group(1)
                            
                    if book_num:
                        link_el = await item.query_selector('a[href*="/book/show/"]')
                        if link_el:
                            b_url = await link_el.get_attribute("href")
                            if b_url and not b_url.startswith('http'): b_url = "https://www.goodreads.com" + b_url
                            books_to_scrape.append((book_num, b_url))
                            
                total_pages = 0
                extracted_data["Goodreads Primary Books Number"] = len(books_to_scrape)
                
                for book_num, b_url in books_to_scrape:
                    b_page = await context.new_page()
                    try:
                        await b_page.goto(b_url, wait_until="domcontentloaded", timeout=45000)
                        extracted_pages = 0
                        
                        ld_el = await b_page.query_selector('script[type="application/ld+json"]')
                        if ld_el:
                            try:
                                data = json.loads(await ld_el.inner_text())
                                if isinstance(data, list): data = data[0]
                                if 'numberOfPages' in data: extracted_pages = int(data['numberOfPages'])
                            except: pass
                            
                        if not extracted_pages:
                            p_el = await b_page.query_selector('[data-testid="pagesFormat"]')
                            if p_el:
                                p_match = re.search(r'(\d+)\s*pages', await p_el.inner_text(), re.IGNORECASE)
                                if p_match: extracted_pages = int(p_match.group(1))
                                
                        if not extracted_pages:
                            try:
                                content = await b_page.content()
                                matches = re.findall(r'(\d+)\s*pages', content, re.IGNORECASE)
                                if matches: extracted_pages = int(matches[0])
                            except: pass
                            
                        if extracted_pages > 0:
                            total_pages += extracted_pages
                    except Exception as e:
                        pass
                    finally:
                        await b_page.close()
                        
                extracted_data["Goodreads Primary Books Page Count"] = total_pages
                
            except Exception as e:
                print(f"[Row {row}] [SERIES] Error: {e}")
                raise
            finally:
                await series_page.close()
        
        # 4. Save to Excel
        async with lock:
            print(f"\n[Row {row}] ======== EXTRACTED DATA ========")
            for col_name, val in extracted_data.items():
                if val is not None:
                    # Truncate very long strings (like Synopsis) for the terminal
                    val_str = str(val)
                    if len(val_str) > 80:
                        val_str = val_str[:77] + "..."
                    print(f"[Row {row}] {col_name}: {val_str}")
            print(f"[Row {row}] ====================================\n")
            
            print(f"[Row {row}] [SAVE] Saved Extracted Data to Excel successfully!")
            for col_name, val in extracted_data.items():
                if val is not None and col_name in headers:
                    col_idx = headers[col_name]
                    # Update cell only if the value is somewhat meaningful (e.g., >0 for ints)
                    if isinstance(val, (int, float)) and val <= 0 and col_name in ["Goodreads Primary Books Number", "Goodreads Primary Books Page Count"]:
                        continue
                    ws.cell(row=row, column=col_idx, value=val)
            wb.save(EXCEL_FILE)

async def main():
    print(f"Loading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active # Use active sheet
    
    # Locate headers
    header_row = 1
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val:
            headers[str(val).strip()] = c
            
    # Find rows missing Goodreads Link or Page Count
    rows_to_process = []
    link_col = headers.get('Goodreads Link')
    pages_col = headers.get('Goodreads Primary Books Page Count')
    pub_col = headers.get('Publisher')
    syn_col = headers.get('Synopsis')
    
    if not link_col:
        print("Could not find 'Goodreads Link' column!")
        return
        
    for row in range(2, ws.max_row + 1):
        link = ws.cell(row=row, column=link_col).value
        pages = ws.cell(row=row, column=pages_col).value if pages_col else None
        
        # Missing link OR missing pages (if link exists, we might just be missing pages/details)
        if not link or str(link).lower() == 'nan' or not pages or pages == 0:
            rows_to_process.append(row)
            
    print(f"Found {len(rows_to_process)} missing rows to process.")
    
    # LIMIT TO 10 for testing
    rows_to_process = rows_to_process[:10]
    print(f"Testing the next {len(rows_to_process)} missing rows...")
    
    if not rows_to_process:
        return
        
    user_data_dir = os.path.join(os.path.dirname(EXCEL_FILE), "playwright_goodreads_profile")
    
    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir,
            headless=False,
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        sem = asyncio.Semaphore(5)
        lock = asyncio.Lock()
        
        async def retry_process_row(row, context, sem, ws, wb, headers, lock):
            max_retries = 3
            for attempt in range(1, max_retries + 1):
                try:
                    await process_row(row, context, sem, ws, wb, headers, lock)
                    return
                except Exception as e:
                    print(f"[{row}] Attempt {attempt} failed: {e}")
                    if attempt == max_retries:
                        print(f"[{row}] All {max_retries} attempts failed.")
                    else:
                        await asyncio.sleep(3)
                        
        tasks = [asyncio.create_task(retry_process_row(row, context, sem, ws, wb, headers, lock)) for row in rows_to_process]
        await asyncio.gather(*tasks)
        
        print("\nAll done!")

if __name__ == "__main__":
    asyncio.run(main())
