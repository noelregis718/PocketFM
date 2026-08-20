import openpyxl

file_path = "E:\\Internship\\PocketFM\\Licensing Outreach Exclusion List.xlsx"
wb = openpyxl.load_workbook(file_path)

sheet_name = None
for s in wb.sheetnames:
    if 'consolidated' in s.lower():
        sheet_name = s
        break

if not sheet_name:
    sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]

series_col = 5
gr_col = 6

seen_series = set()
seen_gr_links = set()
rows_to_delete = []

for row_idx in range(1, ws.max_row + 1):
    series_val = ws.cell(row=row_idx, column=series_col).value
    gr_val = ws.cell(row=row_idx, column=gr_col).value
    
    is_duplicate = False
    
    # Check series
    series_name = None
    if series_val and isinstance(series_val, str):
        series_name = series_val.strip().lower()
        if series_name != '' and series_name != 'none' and series_name in seen_series:
            is_duplicate = True
            
    # Check GR link
    gr_link = None
    if gr_val and isinstance(gr_val, str):
        gr_link = gr_val.strip().lower()
        if gr_link != '' and gr_link not in ['na', 'n/a', '-', 'none'] and gr_link in seen_gr_links:
            is_duplicate = True

    if is_duplicate:
        rows_to_delete.append(row_idx)
    else:
        if series_name and series_name != '' and series_name != 'none':
            seen_series.add(series_name)
        if gr_link and gr_link != '' and gr_link not in ['na', 'n/a', '-', 'none']:
            seen_gr_links.add(gr_link)

for row_idx in reversed(rows_to_delete):
    ws.delete_rows(row_idx, 1)

print(f"Deleted {len(rows_to_delete)} duplicate rows based on Series or GR Links.")
wb.save(file_path)
