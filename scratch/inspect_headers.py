import openpyxl

file_path = "E:\\Internship\\PocketFM\\Licensing Outreach Exclusion List.xlsx"
wb = openpyxl.load_workbook(file_path)

sheet_name = None
for s in wb.sheetnames:
    if 'consolidated' in s.lower():
        sheet_name = s
        break

if not sheet_name:
    sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]

print("--- Header Columns ---")
for row_idx in range(1, 5):
    header = []
    for col_idx in range(1, ws.max_column + 1):
        header.append(str(ws.cell(row=row_idx, column=col_idx).value))
    print(f"Row {row_idx}: {header[:15]}")
