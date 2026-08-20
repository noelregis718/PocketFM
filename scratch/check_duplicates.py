import openpyxl
from collections import Counter

file_path = "E:\\Internship\\PocketFM\\Romantasy _ Self Publication Master.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

sheet_name = None
for s in wb.sheetnames:
    if 'picks for licensing' in s.lower():
        sheet_name = s
        break

if not sheet_name:
    sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]

series_col = 2
series_names = []

for row_idx in range(1, ws.max_row + 1):
    cell_value = ws.cell(row=row_idx, column=series_col).value
    if cell_value and isinstance(cell_value, str):
        series_name = cell_value.strip().lower()
        if series_name == 'book series name' or series_name == '':
            continue
        series_names.append(series_name)

counter = Counter(series_names)
duplicates = {name: count for name, count in counter.items() if count > 1}

print(f"Total entries checked: {len(series_names)}")
if not duplicates:
    print("No duplicates found. The list is completely clean!")
else:
    print(f"Found {len(duplicates)} duplicate series names:")
    for name, count in list(duplicates.items())[:10]:
        print(f" - '{name}' appears {count} times")
