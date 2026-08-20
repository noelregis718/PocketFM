import openpyxl

file_path = "E:\\Internship\\PocketFM\\Licensing Outreach Exclusion List.xlsx"
wb = openpyxl.load_workbook(file_path)

sheet_name = None
for s in wb.sheetnames:
    if 'consolidated' in s.lower():
        sheet_name = s
        break

if not sheet_name:
    sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]

series_col = None
gr_col = None

for row_idx in range(1, min(10, ws.max_row + 1)):
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val and isinstance(val, str):
            val_lower = val.lower()
            if not series_col and 'series' in val_lower:
                series_col = col_idx
            if not gr_col and ('gr ' in val_lower or 'goodreads' in val_lower or 'gr_link' in val_lower or 'grlink' in val_lower):
                gr_col = col_idx
    if series_col and gr_col:
        break

if not series_col:
    print("Could not find 'Series' column. Using 6.")
    series_col = 6
if not gr_col:
    print("Could not find 'GR Link' column. Using 12 as fallback.")
    gr_col = 12

print(f"Using column {series_col} for Series, column {gr_col} for GR Links.")

seen_series = set()
seen_gr_links = set()
rows_to_delete = []

for row_idx in range(1, ws.max_row + 1):
    series_val = ws.cell(row=row_idx, column=series_col).value
    gr_val = ws.cell(row=row_idx, column=gr_col).value
    
    is_duplicate = False
    
    series_name = None
    if series_val and isinstance(series_val, str):
        series_name = series_val.strip().lower()
        if 'series' in series_name and len(series_name) < 15 and row_idx < 5:
            continue
        if series_name in seen_series:
            is_duplicate = True
            
    gr_link = None
    if gr_val and isinstance(gr_val, str):
        gr_link = gr_val.strip().lower()
        if 'goodreads' in gr_link and len(gr_link) < 20 and row_idx < 5:
             pass
        elif gr_link != '' and gr_link not in ['na', 'n/a', '-', 'none'] and gr_link in seen_gr_links:
            is_duplicate = True

    if is_duplicate:
        rows_to_delete.append(row_idx)
    else:
        if series_name and series_name != '':
            seen_series.add(series_name)
        if gr_link and gr_link != '' and gr_link not in ['na', 'n/a', '-', 'none']:
            seen_gr_links.add(gr_link)

for row_idx in reversed(rows_to_delete):
    ws.delete_rows(row_idx, 1)

print(f"Deleted {len(rows_to_delete)} duplicate rows based on Series or GR Links.")
wb.save(file_path)
