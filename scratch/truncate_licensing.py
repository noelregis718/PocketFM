import openpyxl

file_path = "E:\\Internship\\PocketFM\\Licensing Outreach Exclusion List.xlsx"
wb = openpyxl.load_workbook(file_path)

sheet_name = None
for s in wb.sheetnames:
    if 'consolidated' in s.lower():
        sheet_name = s
        break

if not sheet_name:
    sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]

# We want to keep up to row 11282.
start_delete = 11283
amount_to_delete = ws.max_row - start_delete + 1

if amount_to_delete > 0:
    ws.delete_rows(start_delete, amount_to_delete)
    print(f"Successfully deleted {amount_to_delete} rows from row {start_delete} to the end. New max row is {ws.max_row}")
else:
    print("No rows needed to be deleted.")

wb.save(file_path)
