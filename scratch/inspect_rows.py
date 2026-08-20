import openpyxl

file_path = "E:\\Internship\\PocketFM\\Romantasy _ Self Publication Master - 1.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

# Find the 'picks for licensing' sheet
sheet_to_keep = None
for s in wb.sheetnames:
    if 'picks for licensing' in s.lower():
        sheet_to_keep = s
        break

if not sheet_to_keep:
    sheet_to_keep = wb.sheetnames[0]

ws = wb[sheet_to_keep]

print(f"Sheet: {sheet_to_keep}")
print("--- First 5 rows ---")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
    print(f"Row {i+1}: {row[:5]} ...")

print("--- Searching for 'perfect match' ---")
found = False
for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    for cell in row:
        if cell and isinstance(cell, str) and 'perfect match' in cell.lower():
            print(f"Found 'perfect match' on row {row_idx}: {row[:5]} ...")
            found = True
            break
    if found:
        break
