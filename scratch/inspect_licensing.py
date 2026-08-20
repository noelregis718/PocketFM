import openpyxl

file_path = "E:\\Internship\\PocketFM\\Licensing Outreach Exclusion List.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

sheet_name = None
for s in wb.sheetnames:
    if 'consolidated' in s.lower():
        sheet_name = s
        break

if not sheet_name:
    sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]

print("--- Inspecting rows around 11282 ---")
start_row = max(1, 11275)
end_row = min(ws.max_row, 11290)

for row_idx in range(start_row, end_row + 1):
    row_vals = [str(cell.value) if cell.value is not None else "" for cell in ws[row_idx]]
    # Print the first few columns
    print(f"Row {row_idx}: {row_vals[:5]}")

print("--- Searching globally for 'monstrous matches' ---")
found = False
for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    for cell in row:
        if cell and isinstance(cell, str) and 'monstrous matches' in cell.lower():
            print(f"Found 'monstrous matches' in cell on row {row_idx}: {row[:5]}")
            found = True
            break

print("--- Searching globally for 'vampire' near row 11282 ---")
for row_idx in range(max(1, 11200), min(ws.max_row + 1, 11300)):
    row_vals = [str(cell.value) if cell.value is not None else "" for cell in ws[row_idx]]
    for cell in row_vals:
         if 'vampire' in cell.lower():
             print(f"Found 'vampire' on row {row_idx}: {row_vals[:5]}")
             break
