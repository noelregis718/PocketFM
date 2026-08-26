import openpyxl
import sys

def modify_excel(file_path):
    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    
    target_sheet = None
    # Find the sheet case-insensitively or loosely
    for sheet_name in wb.sheetnames:
        if "self pub prioritization" in sheet_name.lower() or "self pub" in sheet_name.lower():
            target_sheet = sheet_name
            break
            
    if not target_sheet:
        print("Could not find a sheet matching 'self pub prioritization'.")
        print("Available sheets:", wb.sheetnames)
        sys.exit(1)
        
    print(f"Found target sheet: {target_sheet}")
    
    # Delete all other sheets
    sheets_to_remove = [s for s in wb.sheetnames if s != target_sheet]
    for sheet_name in sheets_to_remove:
        print(f"Removing sheet: {sheet_name}")
        del wb[sheet_name]
        
    print("Saving the workbook...")
    wb.save(file_path)
    print("Done!")

if __name__ == "__main__":
    file_path = r"e:\Internship\PocketFM\CT _ US _ Pipeline Master Sheet.xlsx"
    modify_excel(file_path)
