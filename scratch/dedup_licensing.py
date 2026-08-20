import openpyxl

file_path = "E:\\Internship\\PocketFM\\Licensing Outreach Exclusion List.xlsx"
wb = openpyxl.load_workbook(file_path)

sheet_name = None
for s in wb.sheetnames:
    if 'consolidated' in s.lower():
        sheet_name = s
        break

if not sheet_name:
    sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]

series_col = None
for row_idx in range(1, min(10, ws.max_row + 1)):
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val and isinstance(val, str) and 'series' in val.lower():
            series_col = col_idx
            break
    if series_col:
        break

if not series_col:
    print("Could not find a column with 'series' in the header. Using Column 5 (E) as default.")
    series_col = 5

seen_series = set()
rows_to_delete = []

for row_idx in range(1, ws.max_row + 1):
    cell_value = ws.cell(row=row_idx, column=series_col).value
    if cell_value and isinstance(cell_value, str):
        series_name = cell_value.strip().lower()
        if 'series' in series_name and len(series_name) < 15 and row_idx < 5: 
            # likely a header, ignore
            continue
            
        if series_name in seen_series:
            rows_to_delete.append(row_idx)
        else:
            seen_series.add(series_name)

for row_idx in reversed(rows_to_delete):
    ws.delete_rows(row_idx, 1)

print(f"Deleted {len(rows_to_delete)} duplicate rows based on column {series_col}.")
wb.save(file_path)
