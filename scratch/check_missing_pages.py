import openpyxl

EXCEL_FILE = r'E:\Internship\PocketFM\Agencies_Scraped.xlsx'
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb['Agencies']

header_row = 2
pages_col = None
series_col = None
author_col = None
for c in range(1, ws.max_column + 1):
    val = ws.cell(row=header_row, column=c).value
    if val == 'No. of pages': pages_col = c
    if val == 'Series Title': series_col = c
    if val == 'Author': author_col = c

missing = 0
for row in range(3, ws.max_row + 1):
    pages = ws.cell(row=row, column=pages_col).value
    series = ws.cell(row=row, column=series_col).value
    author = ws.cell(row=row, column=author_col).value
    
    if series and not pages:
        missing += 1
        print(f'- Excel Row {row}: {series} by {author}')

print(f'\nTotal missing: {missing}')
