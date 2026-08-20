import openpyxl

file_path = "E:\\Internship\\PocketFM\\Romantasy _ Self Publication Master - 1.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

sheet_to_keep = None
for s in wb.sheetnames:
    if 'picks for licensing' in s.lower():
        sheet_to_keep = s
        break

if not sheet_to_keep:
    sheet_to_keep = wb.sheetnames[0]

ws = wb[sheet_to_keep]

print("--- Searching for 'perfect match' in column B ---")
for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if len(row) > 1 and isinstance(row[1], str) and 'perfect match' in row[1].lower():
        print(f"Found '{row[1]}' on row {row_idx}")

print("--- Searching globally for 'perfect match' ---")
for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    for cell in row:
        if cell and isinstance(cell, str) and 'perfect match' in cell.lower():
            if row_idx > 100:
                 print(f"Found 'perfect match' in cell on row {row_idx}: {cell}")
