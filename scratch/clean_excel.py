import openpyxl
import sys

def modify_excel(file_path):
    print(f"Loading {file_path}... (This may take a few minutes for large files)", flush=True)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Error loading workbook: {e}", flush=True)
        sys.exit(1)
        
    print("Workbook loaded successfully.", flush=True)
    
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if "self-pub prioritization" in sheet_name.lower() or "self pub prioritization" in sheet_name.lower():
            target_sheet = sheet_name
            break
            
    if not target_sheet:
        print("Could not find a sheet matching 'self pub prioritization'.", flush=True)
        print("Available sheets:", wb.sheetnames, flush=True)
        sys.exit(1)
        
    print(f"Found target sheet: {target_sheet}", flush=True)
    
    # Delete all other sheets
    sheets_to_remove = [s for s in wb.sheetnames if s != target_sheet]
    for sheet_name in sheets_to_remove:
        print(f"Removing sheet: {sheet_name}", flush=True)
        del wb[sheet_name]
        
    # Process the target sheet
    ws = wb[target_sheet]
    print(f"Processing sheet: {target_sheet}", flush=True)
    
    # 1. Remove AutoFilter
    if ws.auto_filter:
        print("Removing AutoFilter...", flush=True)
        ws.auto_filter.ref = None
        
    # 2. Remove Data Validations
    print("Removing Data Validations...", flush=True)
    ws.data_validations.dataValidation = []
    
    print("Saving the workbook... (This may also take a few minutes)", flush=True)
    try:
        wb.save(file_path)
        print("Done! File saved successfully.", flush=True)
    except Exception as e:
        print(f"Error saving workbook: {e}", flush=True)
        sys.exit(1)

if __name__ == "__main__":
    file_path = r"e:\Internship\PocketFM\CT _ US _ Pipeline Master Sheet.xlsx"
    modify_excel(file_path)
