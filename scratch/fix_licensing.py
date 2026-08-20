import openpyxl
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

file_path = "E:\\Internship\\PocketFM\\Licensing Outreach Exclusion List.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

keep_sheet_name = 'consolidated ips'
sheet_names = wb.sheetnames
sheet_to_keep = None

for sheet in sheet_names:
    if 'consolidated' in sheet.lower():
        sheet_to_keep = sheet
        break

if not sheet_to_keep:
    sheet_to_keep = sheet_names[0]

for sheet in sheet_names:
    if sheet != sheet_to_keep:
        del wb[sheet]
        print(f"Deleted sheet: {sheet}")

ws = wb[sheet_to_keep]
wb[sheet_to_keep].protection.sheet = False
if wb.security:
    wb.security.lockStructure = False

target_row = None
for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    found = False
    for cell in row:
        if cell and isinstance(cell, str) and 'monstrous matches' in cell.lower():
            target_row = row_idx
            found = True
            break
    if found:
        break

if not target_row:
    target_row = 11282
    print(f"Could not find 'monstrous matches', falling back to row {target_row}")
else:
    print(f"Found 'monstrous matches' on row {target_row}")

start_delete = 2
end_delete = target_row - 1
amount_to_delete = end_delete - start_delete + 1

if amount_to_delete > 0:
    ws.delete_rows(start_delete, amount_to_delete)
    print(f"Successfully deleted {amount_to_delete} rows from {start_delete} to {end_delete}. New max row is {ws.max_row}")
else:
    print("No rows deleted (target_row is too small).")

wb.save(file_path)
print(f"Finished processing '{file_path}'.")
