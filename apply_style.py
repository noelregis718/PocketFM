import openpyxl

file_path = r'E:\Internship\PocketFM\Merged_Romance_Keywords.xlsx'
print(f'Applying thin rows (15) and wrap text styling to {file_path}...')
try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 15
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.alignment:
                cell.alignment = openpyxl.styles.Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    text_rotation=cell.alignment.text_rotation,
                    wrap_text=True,
                    shrink_to_fit=cell.alignment.shrink_to_fit,
                    indent=cell.alignment.indent
                )
            else:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                
    wb.save(file_path)
    print('Styling applied successfully.')
except Exception as e:
    print(f'Failed to apply styling: {e}')
